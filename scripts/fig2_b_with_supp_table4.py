#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fig2b: Network composition (% of ripples) across frequency, HC vs SZ (single figure, shared y-axis)

Input:
  NN_open_code/data/events_by_network_subject_level.csv
Required columns:
  site, group, subject, freq, network, n_events

Outputs:
  outputs/tables/Fig2b_network_percent_perm_tests.csv
  outputs/figures/Fig2b_group_network_freq_percent_HC_vs_SZ.pdf
"""

import argparse
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from statsmodels.stats.multitest import multipletests


# -------------------------
# Parameters
# -------------------------

FREQ_ORDER = [80, 120, 160, 200, 240]
NET_ORDER = [
    "Default",
    "Frontoparietal",
    "Dorsal Attention",
    "Ventral Attention",
    "Somatomotor",
    "Visual",
    "Limbic",
    "Hippocampus",
]
GROUP_ORDER = ["HC", "SZ"]

GROUP_COLOR = {
    "HC": "#4C72B0",  # blue
    "SZ": "#C44E52",  # red
}


# -------------------------
# Normalizers
# -------------------------

def normalize_group(g: str) -> str:
    s = str(g).strip().upper()
    if s in ("SC", "SCHIZOPHRENIA"):
        return "SZ"
    if s in ("HEALTHY", "CONTROL"):
        return "HC"
    return s

def normalize_network(name: str) -> str:
    s = str(name).strip().replace("-", " ").replace("_", " ")
    s = s.replace("Network", "").replace("Mode", "")
    s = " ".join(s.split())
    key = s.lower().replace(" ", "")

    lut = {
        "default":          "Default",
        "defaultmode":      "Default",
        "dmn":              "Default",

        "frontoparietal":   "Frontoparietal",
        "fpn":              "Frontoparietal",

        "dorsalattention":  "Dorsal Attention",
        "dan":              "Dorsal Attention",

        "ventralattention": "Ventral Attention",
        "van":              "Ventral Attention",

        "somatomotor":      "Somatomotor",
        "smn":              "Somatomotor",

        "visual":           "Visual",
        "vn":               "Visual",

        "limbic":           "Limbic",
        "ln":               "Limbic",

        "hippocampus":      "Hippocampus",
        "hpc":              "Hippocampus",
    }
    if key in lut:
        return lut[key]
    if s in NET_ORDER:
        return s
    return s


# -------------------------
# Permutation test (same style as your code)
# -------------------------

def perm_test_mean_diff(hc_vals, sz_vals, *, n_perm=10000, seed=0):
    """
    Two-sided permutation test:
      obs = mean(SZ) - mean(HC)
      p = proportion(|perm| >= |obs|)
    """
    rng = np.random.default_rng(seed)
    x = np.asarray(hc_vals, float); x = x[np.isfinite(x)]
    y = np.asarray(sz_vals, float); y = y[np.isfinite(y)]
    if (x.size < 2) or (y.size < 2):
        return np.nan, np.nan

    obs = float(np.mean(y) - np.mean(x))
    cat = np.concatenate([x, y])
    nx = x.size
    diffs = np.empty(n_perm, float)
    for i in range(n_perm):
        rng.shuffle(cat)
        diffs[i] = float(np.mean(cat[nx:]) - np.mean(cat[:nx]))

    p = float(np.mean(np.abs(diffs) >= np.abs(obs)))
    p = max(p, 1.0 / float(n_perm))  # avoid exact zero
    return obs, p


# -------------------------
# Main
# -------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default=None)
    ap.add_argument("--input", type=str, default=None,
                    help="CSV path (default: <root>/data/events_by_network_subject_level.csv)")
    ap.add_argument("--n-perm", type=int, default=10000)
    ap.add_argument("--seed", type=int, default=0)
    args = ap.parse_args()

    if args.root is not None:
        root = Path(args.root).expanduser().resolve()
    else:
        try:
            root = Path(__file__).resolve().parents[1]
        except NameError:
            root = Path.cwd().resolve()

    in_csv = Path(args.input).expanduser().resolve() if args.input else (root / "data" / "events_by_network_subject_level.csv")
    out_tab = root / "outputs" / "tables"
    out_fig = root / "outputs" / "figures"
    out_tab.mkdir(parents=True, exist_ok=True)
    out_fig.mkdir(parents=True, exist_ok=True)

    if not in_csv.exists():
        raise FileNotFoundError(f"Input not found: {in_csv}")

    df = pd.read_csv(in_csv)

    need = {"site", "group", "subject", "freq", "network", "n_events"}
    miss = need - set(df.columns)
    if miss:
        raise ValueError(f"Missing required columns: {sorted(miss)}")

    # normalize + filter
    df = df.copy()
    df["group"] = df["group"].map(normalize_group)
    df["network"] = df["network"].map(normalize_network)
    df["freq"] = pd.to_numeric(df["freq"], errors="coerce").astype(int)
    df["n_events"] = pd.to_numeric(df["n_events"], errors="coerce")

    df = df[df["freq"].isin(FREQ_ORDER)].copy()
    df = df[df["group"].isin(GROUP_ORDER)].copy()
    df = df[df["network"].isin(NET_ORDER)].copy()
    df = df.dropna(subset=["n_events", "freq", "subject"])

    # pooled across sites: group×subject×freq×network
    df = (df.groupby(["group", "subject", "freq", "network"], as_index=False)["n_events"].sum())

    # subject×freq total -> pct
    tot = (df.groupby(["group", "subject", "freq"], as_index=False)["n_events"]
             .sum()
             .rename(columns={"n_events": "sum_events"}))
    dfp = df.merge(tot, on=["group", "subject", "freq"], how="left")
    dfp.loc[dfp["sum_events"] <= 0, "sum_events"] = np.nan
    dfp["pct"] = (dfp["n_events"] / dfp["sum_events"]) * 100.0

    # group×freq×network mean±SE (save)
    rows = []
    for (grp, freq, net), sub in dfp.groupby(["group", "freq", "network"], dropna=False):
        vals = pd.to_numeric(sub["pct"], errors="coerce").dropna().to_numpy(float)
        if vals.size == 0:
            continue
        mean = float(vals.mean())
        sd = float(vals.std(ddof=1)) if vals.size > 1 else np.nan
        se = float(sd / np.sqrt(vals.size)) if vals.size > 1 else np.nan
        rows.append(dict(group=grp, freq=int(freq), network=net, n=int(vals.size), mean=mean, se=se))
    df_meanse = pd.DataFrame(rows)

    # HC vs SZ permutation tests (save)
    perm_rows = []
    for (net, freq), g in dfp.groupby(["network", "freq"], dropna=False):
        hc = pd.to_numeric(g.loc[g["group"] == "HC", "pct"], errors="coerce").dropna().to_numpy(float)
        sz = pd.to_numeric(g.loc[g["group"] == "SZ", "pct"], errors="coerce").dropna().to_numpy(float)

        mean_hc = float(np.mean(hc)) if hc.size else np.nan
        mean_sz = float(np.mean(sz)) if sz.size else np.nan
        sd_hc = float(np.std(hc, ddof=1)) if hc.size > 1 else (0.0 if hc.size == 1 else np.nan)
        sd_sz = float(np.std(sz, ddof=1)) if sz.size > 1 else (0.0 if sz.size == 1 else np.nan)

        if (hc.size >= 2) and (sz.size >= 2):
            obs, p = perm_test_mean_diff(hc, sz, n_perm=args.n_perm, seed=args.seed)
        else:
            obs, p = np.nan, np.nan

        perm_rows.append(dict(
            network=str(net), freq=int(freq),
            n_HC=int(hc.size), n_SZ=int(sz.size),
            mean_HC=mean_hc, mean_SZ=mean_sz,
            sd_HC=sd_hc, sd_SZ=sd_sz,
            observed_diff_SZ_minus_HC=float(obs) if np.isfinite(obs) else np.nan,
            p_value=float(p) if np.isfinite(p) else np.nan,
        ))
    df_perm = pd.DataFrame(perm_rows)
    df_perm["q_fdr"] = np.nan
    mask = df_perm["p_value"].notna()
    if mask.any():
        df_perm.loc[mask, "q_fdr"] = multipletests(df_perm.loc[mask, "p_value"], method="fdr_bh")[1]

    # -------------------------
    # Supplementary Table S4 (template-based): Network-specific composition (% of ripples)
    # -------------------------

    # Build mean & SD table (subject-level % already in dfp)
    stat_rows = []
    for net in NET_ORDER:
        for freq in FREQ_ORDER:
            sub = dfp[(dfp["network"] == net) & (dfp["freq"] == freq)]
            hc = sub.loc[sub["group"] == "HC", "pct"].astype(float)
            sz = sub.loc[sub["group"] == "SZ", "pct"].astype(float)

            mean_hc = float(hc.mean()) if hc.size else np.nan
            mean_sz = float(sz.mean()) if sz.size else np.nan
            sd_hc = float(hc.std(ddof=1)) if hc.size > 1 else (0.0 if hc.size == 1 else np.nan)
            sd_sz = float(sz.std(ddof=1)) if sz.size > 1 else (0.0 if sz.size == 1 else np.nan)

            # fetch permutation stats
            row_p = df_perm[(df_perm["network"] == net) & (df_perm["freq"] == freq)]
            p_raw = float(row_p["p_value"].iloc[0]) if len(row_p) else np.nan
            q_fdr = float(row_p["q_fdr"].iloc[0]) if len(row_p) else np.nan

            stat_rows.append(dict(
                network=net,
                freq=int(freq),
                mean_HC=mean_hc,
                mean_SZ=mean_sz,
                sd_HC=sd_hc,
                sd_SZ=sd_sz,
                diff_SZ_minus_HC=(mean_sz - mean_hc) if (np.isfinite(mean_hc) and np.isfinite(mean_sz)) else np.nan,
                p_value=p_raw,
                q_fdr=q_fdr,
            ))

    df_s4 = pd.DataFrame(stat_rows)

    # Save a tidy CSV (useful for downstream checks)
    out_csv_s4 = out_tab / "Table_S4_network_composition.csv"
    df_s4.to_csv(out_csv_s4, index=False)

    # Write XLSX following the provided template (supp_table4.xlsx)
    def _find_template_s4(root: Path) -> Path:
        candidates = [
            root / "templates" / "supp_table4.xlsx",
            root / "template" / "supp_table4.xlsx",
            root / "supp_tables" / "supp_table4.xlsx",
            Path(__file__).resolve().parent / "supp_table4.xlsx",
            Path(__file__).resolve().parent.parent / "supp_table4.xlsx",
        ]
        for c in candidates:
            if c.exists():
                return c
        raise FileNotFoundError("supp_table4.xlsx template was not found in expected locations.")

    def _write_s4_xlsx_from_template(df_s4: pd.DataFrame, template_path: Path, out_path: Path) -> None:
        # Map canonical names to the abbreviations used in the manuscript table.
        lut_abbrev = {
            "Default": "DMN",
            "Frontoparietal": "FPN",
            "Dorsal Attention": "DAN",
            "Ventral Attention": "VAN",
            "Somatomotor": "SMN",
            "Visual": "VN",
            "Limbic": "LM",
            "Hippocampus": "Hippocampus",
        }

        wb = openpyxl.load_workbook(template_path)
        ws = wb[wb.sheetnames[0]]

        # Clear existing data rows (keep headers & caption). Template uses rows 4..43 for data.
        for r in range(4, 44):
            for c in range(1, 10):
                ws.cell(r, c).value = None

        r0 = 4
        for net in NET_ORDER:
            net_disp = lut_abbrev.get(net, str(net))
            for i, freq in enumerate(FREQ_ORDER):
                row = df_s4[(df_s4["network"] == net) & (df_s4["freq"] == int(freq))]
                if row.empty:
                    vals = dict(mean_HC=np.nan, mean_SZ=np.nan, sd_HC=np.nan, sd_SZ=np.nan,
                                diff_SZ_minus_HC=np.nan, p_value=np.nan, q_fdr=np.nan)
                else:
                    vals = row.iloc[0].to_dict()

                rr = r0
                # Network label only on the first row of each block (as in the template)
                ws.cell(rr, 1).value = net_disp if i == 0 else None
                ws.cell(rr, 2).value = int(freq)

                # Numeric columns (keep 3 decimals for means/sds/diff; p/q in 4 decimals)
                def _set_num(cell, x, nd=3):
                    if x is None or (isinstance(x, float) and (not np.isfinite(x))):
                        cell.value = None
                    else:
                        cell.value = round(float(x), nd)

                _set_num(ws.cell(rr, 3), vals.get("mean_HC"), nd=3)
                _set_num(ws.cell(rr, 4), vals.get("mean_SZ"), nd=3)
                _set_num(ws.cell(rr, 5), vals.get("sd_HC"), nd=3)
                _set_num(ws.cell(rr, 6), vals.get("sd_SZ"), nd=3)
                _set_num(ws.cell(rr, 7), vals.get("diff_SZ_minus_HC"), nd=3)

                # p and q
                _set_num(ws.cell(rr, 8), vals.get("p_value"), nd=4)
                _set_num(ws.cell(rr, 9), vals.get("q_fdr"), nd=4)

                r0 += 1

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)

    try:
        template_s4 = _find_template_s4(root)
        out_xlsx_s4 = out_tab / "Table_S4_network_composition.xlsx"
        _write_s4_xlsx_from_template(df_s4, template_s4, out_xlsx_s4)
    except Exception as e:
        print("[WARN] Supplementary Table S4 (xlsx) was not written:", repr(e))
        out_xlsx_s4 = None

    # -------------------------
    # Plot (single figure, shared y-axis)
    # -------------------------

    # shared y-axis max across all networks/ freqs/ groups
    if df_meanse.empty:
        raise RuntimeError("df_meanse is empty: check input CSV contents and network labels.")
    y_max = float((df_meanse["mean"] + df_meanse["se"]).max() * 1.15)

    fig, axes = plt.subplots(2, 4, figsize=(13, 6), sharey=True)
    axes = np.atleast_1d(axes).ravel()

    width = 0.36
    x = np.arange(len(FREQ_ORDER))

    for i, net in enumerate(NET_ORDER):
        ax = axes[i]
        sub = df_meanse[df_meanse["network"] == net].copy()
        if sub.empty:
            ax.axis("off")
            continue

        for gi, grp in enumerate(GROUP_ORDER):
            gsub = sub[sub["group"] == grp].set_index("freq").reindex(FREQ_ORDER)
            means = gsub["mean"].to_numpy(float)
            ses   = gsub["se"].to_numpy(float)

            offset = (-width/2) if gi == 0 else (width/2)
            ax.bar(
                x + offset,
                means,
                width=width,
                yerr=ses,
                capsize=2,
                color=GROUP_COLOR.get(grp),
                edgecolor="black",
                linewidth=0.5,
                label=grp if i == 0 else None,
            )

        ax.set_title(net, fontsize=11)
        ax.set_xticks(x)
        ax.set_xticklabels([str(f) for f in FREQ_ORDER])
        ax.set_ylim(0, y_max)
        if i % 4 == 0:
            ax.set_ylabel("Percent of ripples (%), mean ± SEM")

    # shared legend
    handles = [
        plt.Rectangle((0, 0), 1, 1, color=GROUP_COLOR[g], ec="black")
        for g in GROUP_ORDER
    ]
    fig.legend(handles, GROUP_ORDER, loc="upper right", frameon=False)

    fig.suptitle("Fig2b: Network composition of ripple events (HC vs SZ)", y=1.02, fontsize=13)
    fig.tight_layout()

    out_pdf = out_fig / "Fig2b_group_network_freq_percent_HC_vs_SZ.pdf"
    fig.savefig(out_pdf, bbox_inches="tight")
    plt.close(fig)

    print("[OK] Written:")
    
    print(" -", out_csv_s4)
    if out_xlsx_s4 is not None:
        print(" -", out_xlsx_s4)
    print(" -", out_pdf)


if __name__ == "__main__":
    main()
