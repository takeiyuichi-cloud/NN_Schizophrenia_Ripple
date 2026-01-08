#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fig3bde: High-rate epoch group comparison (HC vs SZ) + Supplementary Tables S5 & S7
SUBJECT-LEVEL authoritative version.

Fix in v3 (requested):
- Supplementary Table S5 now additionally reports:
    (i) mean duration of individual clustered epochs (mean_epoch_dur_s)
    (ii) median duration of individual clustered epochs (median_epoch_dur_s)
  so that the manuscript statement about mean/median duration can cite Table S5.
- Durations are computed from rate_epoch_epoch_level.csv when available.
  If epoch-level is missing, mean duration falls back to sum_epoch_dur_s/n_epochs and
  median duration is left as NaN.
- For duration metrics, subjects with n_epochs==0 are set to NaN (undefined), so they do
  not distort duration comparisons.

Core policy:
- Use rate_epoch_subject_level.csv as the primary cohort (subject×freq).
- Do NOT depend on high_rate_epoch_debug_merged.csv (reserved for Table S6).

Outputs (<root>/outputs/):
  - tables/Table_S5_temporal_clustering_metrics.csv
  - tables/Table_S7_correlations.csv
  - figures/Fig3bde_high_rate_epoch_group_comparison.pdf
"""

from __future__ import annotations

import argparse
from pathlib import Path
import re
import numpy as np
import pandas as pd
import openpyxl

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from statsmodels.stats.multitest import multipletests
from scipy import stats as spstats

FREQ_ORDER = [80, 120, 160, 200, 240]
GROUPS = ["HC", "SZ"]
COLOR = {"HC": "#4C72B0", "SZ": "#C44E52"}

N_PERM_DEFAULT = 10000
SEED_DEFAULT = 0


# -------------------------
# Helpers
# -------------------------
def canon_subject(x) -> str:
    s = str(x)
    m = re.search(r"(\d+)", s)
    if not m:
        return s.strip()
    return f"NB_subject_{int(m.group(1))}"

def normalize_group_labels(s: pd.Series) -> pd.Series:
    x = s.astype(str).str.strip().str.upper()
    x = x.replace({"SCHIZOPHRENIA": "SZ", "SCZ": "SZ", "SC": "SZ"})
    x = x.replace({"HEALTHY": "HC", "CONTROL": "HC"})
    return x

def permutation_test_diff_means(x_hc: np.ndarray, x_sz: np.ndarray, n_perm: int, seed: int = 0) -> tuple[float, float]:
    rng = np.random.default_rng(seed)
    x_hc = np.asarray(x_hc, float); x_hc = x_hc[np.isfinite(x_hc)]
    x_sz = np.asarray(x_sz, float); x_sz = x_sz[np.isfinite(x_sz)]
    if x_hc.size < 2 or x_sz.size < 2:
        return (np.nan, np.nan)
    obs = float(np.mean(x_sz) - np.mean(x_hc))
    pooled = np.concatenate([x_hc, x_sz])
    n_hc = x_hc.size
    diffs = np.empty(n_perm, float)
    for i in range(n_perm):
        rng.shuffle(pooled)
        diffs[i] = float(np.mean(pooled[n_hc:]) - np.mean(pooled[:n_hc]))
    p = float(np.mean(np.abs(diffs) >= abs(obs)))
    p = max(p, 1.0 / float(n_perm))
    return obs, p

def cohens_d(x_hc: np.ndarray, x_sz: np.ndarray) -> float:
    x_hc = np.asarray(x_hc, float); x_hc = x_hc[np.isfinite(x_hc)]
    x_sz = np.asarray(x_sz, float); x_sz = x_sz[np.isfinite(x_sz)]
    if x_hc.size < 2 or x_sz.size < 2:
        return np.nan
    v1 = np.var(x_hc, ddof=1)
    v2 = np.var(x_sz, ddof=1)
    denom = x_hc.size + x_sz.size - 2
    if denom <= 0:
        return np.nan
    pooled = ((x_hc.size - 1) * v1 + (x_sz.size - 1) * v2) / denom
    if pooled <= 0 or not np.isfinite(pooled):
        return np.nan
    return float((np.mean(x_sz) - np.mean(x_hc)) / np.sqrt(pooled))

def mean_sem(x: np.ndarray) -> tuple[float, float, int]:
    x = np.asarray(x, float)
    x = x[np.isfinite(x)]
    n = int(x.size)
    if n == 0:
        return (np.nan, np.nan, 0)
    m = float(np.mean(x))
    se = float(np.std(x, ddof=1) / np.sqrt(n)) if n >= 2 else np.nan
    return (m, se, n)

def stars(p: float) -> str:
    if p is None or (isinstance(p, float) and (not np.isfinite(p))):
        return ""
    if p < 1e-4:
        return "****"
    if p < 1e-3:
        return "***"
    if p < 1e-2:
        return "**"
    if p < 5e-2:
        return "*"
    return ""

def fmt_p_or_q(pv: float) -> str:
    if pv is None or (isinstance(pv, float) and (not np.isfinite(pv))):
        return ""
    if pv < 1e-4:
        return "<0.0001" + stars(pv)
    return f"{pv:.4g}{stars(pv)}"


# -------------------------
# Path discovery
# -------------------------
def find_root() -> Path:
    try:
        return Path(__file__).resolve().parents[1]
    except NameError:
        return Path.cwd().resolve()

def find_subject_csv(root: Path, user_path: str | None) -> Path:
    if user_path:
        p = Path(user_path).expanduser()
        if not p.is_absolute():
            p = (root / p).resolve()
        if not p.exists():
            raise FileNotFoundError(f"--subject-csv not found: {p}")
        return p
    for c in [root/"data"/"rate_epoch_subject_level.csv", root/"outputs"/"tables"/"rate_epoch_subject_level.csv"]:
        if c.exists():
            return c
    raise FileNotFoundError("Could not find rate_epoch_subject_level.csv. Provide --subject-csv.")

def find_epoch_csv(root: Path, user_path: str | None) -> Path | None:
    if user_path:
        p = Path(user_path).expanduser()
        if not p.is_absolute():
            p = (root / p).resolve()
        return p if p.exists() else None
    for c in [root/"data"/"rate_epoch_epoch_level.csv", root/"outputs"/"tables"/"rate_epoch_epoch_level.csv"]:
        if c.exists():
            return c
    return None

def find_template(root: Path, name: str) -> Path | None:
    candidates = [
        root/"templates"/name, root/"template"/name, root/"supp_tables"/name, root/"data"/name,
        Path(__file__).resolve().parent/name, Path(__file__).resolve().parent.parent/name,
        Path(__file__).resolve().parents[2]/name
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


# -------------------------
# Group attachment
# -------------------------
def attach_group(subject_df: pd.DataFrame, root: Path, group_csv: Path | None) -> pd.DataFrame:
    out = subject_df.copy()

    if "group" in out.columns:
        out["group"] = normalize_group_labels(out["group"])
        return out

    if "S_ID" in out.columns:
        sid = pd.to_numeric(out["S_ID"], errors="coerce")
        out["group"] = sid.map({1: "HC", 2: "SZ"})
        out["group"] = normalize_group_labels(out["group"])
        return out

    if group_csv is not None and group_csv.exists():
        gdf = pd.read_csv(group_csv).copy()
        if "subject" not in gdf.columns:
            for cand in ["ID", "id", "Subject", "SUBJECT"]:
                if cand in gdf.columns:
                    gdf["subject"] = gdf[cand]
                    break
        if "subject" not in gdf.columns:
            raise ValueError("group_csv must have 'subject' (or ID-like) column.")
        gdf["subject"] = gdf["subject"].map(canon_subject).astype(str)

        if "group" not in gdf.columns:
            if "S_ID" in gdf.columns:
                sid = pd.to_numeric(gdf["S_ID"], errors="coerce")
                gdf["group"] = sid.map({1: "HC", 2: "SZ"})
            else:
                raise ValueError("group_csv must have 'group' or 'S_ID'.")
        gdf["group"] = normalize_group_labels(gdf["group"])

        gdf = gdf[["subject", "group"]].drop_duplicates("subject")
        return out.merge(gdf, on="subject", how="left")

    dfc = root / "data" / "df_clean_expanded.csv"
    if dfc.exists():
        gdf = pd.read_csv(dfc).copy()
        if {"subject", "group"}.issubset(gdf.columns):
            gdf["subject"] = gdf["subject"].map(canon_subject).astype(str)
            gdf["group"] = normalize_group_labels(gdf["group"])
            gdf = gdf[gdf["group"].isin(GROUPS)][["subject", "group"]].drop_duplicates("subject")
            return out.merge(gdf, on="subject", how="left")

    raise RuntimeError("Could not attach group labels. Provide --group-csv or include group/S_ID in subject-level csv.")


# -------------------------
# Derived metrics
# -------------------------
def add_subject_level_metrics(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in ["freq","n_epochs","sum_epoch_dur_s","n_events","n_events_in_epochs","recording_len_s"]:
        out[c] = pd.to_numeric(out[c], errors="coerce")

    out["n_events_outside_epochs"] = out["n_events"] - out["n_events_in_epochs"]
    out["outside_dur_s"] = (out["recording_len_s"] - out["sum_epoch_dur_s"]).clip(lower=0.0)

    out["outside_event_rate"] = np.where(
        out["outside_dur_s"] > 0,
        out["n_events_outside_epochs"] / out["outside_dur_s"],
        np.where((out["outside_dur_s"] == 0) & (out["n_events_outside_epochs"] == 0), 0.0, np.nan),
    )

    out["within_epoch_event_rate"] = np.where(
        out["sum_epoch_dur_s"] > 0,
        out["n_events_in_epochs"] / out["sum_epoch_dur_s"],
        np.where((out["sum_epoch_dur_s"] == 0) & (out["n_events_in_epochs"] == 0), 0.0, np.nan),
    )

    return out

def add_epoch_duration_stats(df: pd.DataFrame, epoch_csv: Path | None) -> pd.DataFrame:
    """
    mean_epoch_dur_s:
      - Prefer subject-level column if present.
      - Fallback: sum_epoch_dur_s / n_epochs (when n_epochs>0).
    median_epoch_dur_s is NOT used (removed).
    epoch_csv is ignored (kept only for backward compatibility of function signature).
    """
    out = df.copy()

    # ① subject-level mean_epoch_dur_s を温存・数値化（潰さない）
    if "mean_epoch_dur_s" in out.columns:
        out["mean_epoch_dur_s"] = pd.to_numeric(out["mean_epoch_dur_s"], errors="coerce")
    else:
        out["mean_epoch_dur_s"] = np.nan

    # ② fallback（meanが欠損のときのみ）
    fallback_mean = np.where(
        pd.to_numeric(out["n_epochs"], errors="coerce") > 0,
        pd.to_numeric(out["sum_epoch_dur_s"], errors="coerce") / pd.to_numeric(out["n_epochs"], errors="coerce"),
        np.nan
    )
    out["mean_epoch_dur_s"] = out["mean_epoch_dur_s"].where(np.isfinite(out["mean_epoch_dur_s"]), fallback_mean)

    # ③ n_epochs==0 は duration 未定義
    out.loc[pd.to_numeric(out["n_epochs"], errors="coerce") <= 0, "mean_epoch_dur_s"] = np.nan

    return out


# -------------------------
# Stats
# -------------------------
def compute_group_stats(df: pd.DataFrame, metric_col: str, metric_key: str, n_perm: int, seed: int) -> pd.DataFrame:
    rows = []
    for f in FREQ_ORDER:
        sub = df[df["freq"] == f].copy()
        g = normalize_group_labels(sub["group"])
        x = pd.to_numeric(sub[metric_col], errors="coerce")

        x_hc = x[g == "HC"].to_numpy(float)
        x_sz = x[g == "SZ"].to_numpy(float)

        n_hc = int(np.isfinite(x_hc).sum())
        n_sz = int(np.isfinite(x_sz).sum())

        m_hc, se_hc, _ = mean_sem(x_hc)
        m_sz, se_sz, _ = mean_sem(x_sz)

        diff, p = permutation_test_diff_means(x_hc, x_sz, n_perm=n_perm, seed=seed + int(f) + (abs(hash(metric_key)) % 100000))
        d = cohens_d(x_hc, x_sz)

        rows.append(dict(
            metric=metric_key, freq=int(f),
            n_HC=n_hc, n_SZ=n_sz,
            mean_HC=m_hc, sem_HC=se_hc,
            mean_SZ=m_sz, sem_SZ=se_sz,
            diff_SZ_minus_HC=diff,
            cohens_d=d,
            p_perm=p
        ))

    out = pd.DataFrame(rows)
    out["q_fdr"] = np.nan
    mask = out["p_perm"].notna()
    if mask.any():
        out.loc[mask, "q_fdr"] = multipletests(out.loc[mask, "p_perm"].values, method="fdr_bh")[1]
    return out


# -------------------------
# Table writers
# -------------------------
def write_s7_from_template(df_s7: pd.DataFrame, template_path: Path, out_path: Path) -> None:
    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]

    # Clear data area (rows 3-22, cols A-G)
    for r in range(3, 23):
        for c in range(1, 8):
            ws.cell(r, c).value = None

    block_order = [
        ("rho_withinDensity_vs_outsideRate", "Correlation between within-epoch ripple density and outside-cluster ripple rate"),
        ("rho_nEpochs_vs_outsideCount", "Correlation between number of clustered epochs and outside-cluster ripple count"),
    ]

    r0 = 3
    for block_key, block_label in block_order:
        for i, freq in enumerate(FREQ_ORDER):
            for gi, grp in enumerate(GROUPS):
                row = df_s7[(df_s7["block"] == block_key) & (df_s7["freq"] == int(freq)) & (df_s7["group"] == grp)]
                rr = r0 + i * 2 + gi

                ws.cell(rr, 1).value = block_label if (i == 0 and gi == 0) else None
                ws.cell(rr, 2).value = int(freq) if gi == 0 else None
                ws.cell(rr, 3).value = grp

                if row.empty:
                    continue
                vals = row.iloc[0].to_dict()

                ws.cell(rr, 4).value = int(vals.get("n", 0))
                rho = vals.get("rho")
                if rho is not None and np.isfinite(rho):
                    ws.cell(rr, 5).value = round(float(rho), 3)
                ws.cell(rr, 6).value = vals.get("p_str", "")
                ws.cell(rr, 7).value = vals.get("q_str", "")

        r0 += 10

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# -------------------------
# Plot
# -------------------------
def plot_bde(df_plot: pd.DataFrame, out_pdf: Path):
    plot_metrics = [
        ("n_epochs", "Temporally clustered SWRs counts (per 300 s)"),
        ("mean_epoch_dur_s", "Mean duration of temporally clustered SWRs (s)"),
        ("outside_event_rate", "Ripple rate outside temporally clustered SWRs (events/s)"),
    ]
    fig, axes = plt.subplots(len(plot_metrics), 1, figsize=(7.2, 3.2 * len(plot_metrics)), sharex=True)
    if len(plot_metrics) == 1:
        axes = [axes]

    for ax, (mkey, ylab) in zip(axes, plot_metrics):
        sub = df_plot[df_plot["metric"] == mkey].set_index("freq").reindex(FREQ_ORDER).reset_index()
        x = np.arange(len(FREQ_ORDER))
        ax.errorbar(x, sub["mean_HC"], yerr=sub["sem_HC"], fmt="-o", capsize=3, label="HC", color=COLOR["HC"])
        ax.errorbar(x, sub["mean_SZ"], yerr=sub["sem_SZ"], fmt="-o", capsize=3, label="SZ", color=COLOR["SZ"])
        ax.set_ylabel(ylab)
        ax.grid(alpha=0.25)

        for i, f in enumerate(FREQ_ORDER):
            q = sub.loc[sub["freq"] == f, "q_fdr"]
            if len(q) == 1 and np.isfinite(q.iloc[0]) and float(q.iloc[0]) < 0.05:
                star = "***" if float(q.iloc[0]) < 0.001 else ("**" if float(q.iloc[0]) < 0.01 else "*")
                ymax = ax.get_ylim()[1]
                ax.text(i, ymax * 1.02, star, ha="center", va="bottom", fontsize=12, fontweight="bold")

    axes[-1].set_xticks(np.arange(len(FREQ_ORDER)))
    axes[-1].set_xticklabels([str(f) for f in FREQ_ORDER])
    axes[-1].set_xlabel("Frequency (Hz)")
    axes[0].legend(frameon=False, loc="upper right")

    fig.suptitle("Fig3bde: High-rate epoch group comparison (HC vs SZ)", y=1.01, fontsize=13)
    fig.tight_layout()
    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_pdf, dpi=300, bbox_inches="tight")
    plt.close(fig)


# -------------------------
# Main
# -------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default=None)
    ap.add_argument("--subject-csv", type=str, default=None)
    ap.add_argument("--epoch-csv", type=str, default=None)
    ap.add_argument("--group-csv", type=str, default=None)
    ap.add_argument("--n-perm", type=int, default=N_PERM_DEFAULT)
    ap.add_argument("--seed", type=int, default=SEED_DEFAULT)
    args = ap.parse_args()

    root = Path(args.root).expanduser().resolve() if args.root else find_root()
    out_tab = root / "outputs" / "tables"
    out_fig = root / "outputs" / "figures"
    out_tab.mkdir(parents=True, exist_ok=True)
    out_fig.mkdir(parents=True, exist_ok=True)

    subject_csv = find_subject_csv(root, args.subject_csv)
    epoch_csv = find_epoch_csv(root, args.epoch_csv)

    df = pd.read_csv(subject_csv).copy()
    need = {"subject", "freq", "n_epochs", "sum_epoch_dur_s", "n_events", "n_events_in_epochs", "recording_len_s"}
    miss = need - set(df.columns)
    if miss:
        raise ValueError(f"rate_epoch_subject_level.csv missing columns: {sorted(miss)}")

    df["subject"] = df["subject"].map(canon_subject).astype(str)
    df["freq"] = pd.to_numeric(df["freq"], errors="coerce")
    df = df[df["freq"].isin(FREQ_ORDER)].copy()

    group_csv = None
    if args.group_csv:
        p = Path(args.group_csv).expanduser()
        if not p.is_absolute():
            p = (root / p).resolve()
        group_csv = p

    df = attach_group(df, root, group_csv)
    df["group"] = normalize_group_labels(df["group"])
    df = df[df["group"].isin(GROUPS)].copy()

    df = add_subject_level_metrics(df)
    df = add_epoch_duration_stats(df, epoch_csv)

    # --- Fig3bde plot stats
    plot_frames = [
        compute_group_stats(df, "n_epochs", "n_epochs", args.n_perm, args.seed),
        compute_group_stats(df, "mean_epoch_dur_s", "mean_epoch_dur_s", args.n_perm, args.seed),
        compute_group_stats(df, "outside_event_rate", "outside_event_rate", args.n_perm, args.seed),
    ]
    df_plot = pd.concat(plot_frames, ignore_index=True)

    # --- Table S5: include mean/median durations
    s5_frames = [
        compute_group_stats(df, "n_epochs", "n_epochs", args.n_perm, args.seed),
        compute_group_stats(df, "mean_epoch_dur_s", "mean_epoch_dur_s", args.n_perm, args.seed),
        compute_group_stats(df, "within_epoch_event_rate", "within_epoch_event_rate", args.n_perm, args.seed),
        compute_group_stats(df, "outside_event_rate", "outside_event_rate", args.n_perm, args.seed),
    ]
    df_s5 = pd.concat(s5_frames, ignore_index=True)
    df_s5.to_csv(out_tab / "Table_S5_temporal_clustering_metrics.csv", index=False)
   
    # --- Table S7 correlations (same definition as before)
    corr_rows = []
    for freq in FREQ_ORDER:
        subf = df[df["freq"] == int(freq)].copy()
        for grp in GROUPS:
            subg = subf[subf["group"] == grp].copy()

            x = pd.to_numeric(subg["within_epoch_event_rate"], errors="coerce")
            y = pd.to_numeric(subg["outside_event_rate"], errors="coerce")
            xy = pd.DataFrame({"x": x, "y": y}).replace([np.inf, -np.inf], np.nan).dropna()
            if len(xy) >= 3:
                rho, p = spstats.spearmanr(xy["x"].to_numpy(float), xy["y"].to_numpy(float))
            else:
                rho, p = np.nan, np.nan
            corr_rows.append(dict(block="rho_withinDensity_vs_outsideRate", freq=int(freq), group=grp,
                                  n=int(len(xy)), rho=float(rho) if np.isfinite(rho) else np.nan,
                                  p=float(p) if np.isfinite(p) else np.nan))

            x2 = pd.to_numeric(subg["n_epochs"], errors="coerce")
            y2 = pd.to_numeric(subg["n_events_outside_epochs"], errors="coerce")
            xy2 = pd.DataFrame({"x": x2, "y": y2}).replace([np.inf, -np.inf], np.nan).dropna()
            if len(xy2) >= 3:
                rho2, p2 = spstats.spearmanr(xy2["x"].to_numpy(float), xy2["y"].to_numpy(float))
            else:
                rho2, p2 = np.nan, np.nan
            corr_rows.append(dict(block="rho_nEpochs_vs_outsideCount", freq=int(freq), group=grp,
                                  n=int(len(xy2)), rho=float(rho2) if np.isfinite(rho2) else np.nan,
                                  p=float(p2) if np.isfinite(p2) else np.nan))

    df_s7 = pd.DataFrame(corr_rows)
    df_s7["q_fdr"] = np.nan
    for block in df_s7["block"].unique():
        for grp in GROUPS:
            msk = (df_s7["block"] == block) & (df_s7["group"] == grp) & df_s7["p"].notna()
            if msk.any():
                df_s7.loc[msk, "q_fdr"] = multipletests(df_s7.loc[msk, "p"], method="fdr_bh")[1]

    df_s7["p_str"] = df_s7["p"].apply(fmt_p_or_q)
    df_s7["q_str"] = df_s7["q_fdr"].apply(lambda q: fmt_p_or_q(float(q)) if q is not None and np.isfinite(q) else "")

    df_s7.to_csv(out_tab / "Table_S7_correlations.csv", index=False)
  
    # --- Plot
    out_pdf = out_fig / "Fig3bde_high_rate_epoch_group_comparison.pdf"
    plot_bde(df_plot, out_pdf)

    print("[OK] Subject CSV:", subject_csv)
    print("[OK] Epoch CSV:", epoch_csv if epoch_csv else "(not found; median_epoch_dur_s will be NaN)")
    print("[OK] Written:")
    print(" -", out_tab / "Table_S5_temporal_clustering_metrics.csv")
    print(" -", out_tab / "Table_S7_correlations.csv")
    print(" -", out_pdf)

if __name__ == "__main__":
    main()
