#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Supplementary Table S11 (template-based; recalculated, NOT using models_results_long.csv):
Clinical associations with ripple clustering outcomes (SZ only)

This version matches the manuscript text:
  (i) Number of high-rate ripple epochs (count)                 -> sum(n_epochs) across 80–240 Hz
  (ii) Ripple events outside high-rate epochs (count)           -> sum(n_events_outside_epochs) across 80–240 Hz
  (iii) Mean ripple density within high-rate epochs (events/s)  -> sum(n_events_in_epochs) / sum(sum_epoch_dur_s) across 80–240 Hz

Models (one predictor at a time; covariate-adjusted; SZ only):
  - Count outcomes: Negative Binomial GLM (log link) + C(site_public), offset=log(minutes_sum), HC3
  - Density outcome: OLS + C(site_public), HC3
  - All models additionally adjust for total ripple load (events_sum) and covariates.

Predictors (any subset present in fig5a_source_public.csv is used):
  - PANSS_positive, PANSS_negative, PANSS_pasological, GAF

Inputs:
  <root>/data/fig5a_source_public.csv      (SAFE)
  <root>/data/fig5a_id_map_private.csv     (PRIVATE; subject <-> anon_id)
  <root>/data/rate_epoch_subject_level.csv (from sliding-window pipeline)

Outputs (<root>/outputs/tables/):
  - Table_S11_clinical_associations_with_cluster_metrics.csv
"""

from __future__ import annotations
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

import statsmodels.api as sm
import statsmodels.formula.api as smf
from statsmodels.stats.multitest import multipletests

FREQS_USE = [80, 120, 160, 200, 240]
COVARS = ["age", "sex", "JART", "sleepiness_pre", "antipsychotics"]
PRED_MAP = [
    ("PANSS_positive", "PANSS POS"),
    ("PANSS_negative", "PANSS NEG"),
    ("PANSS_pasological", "PANSS GEN"),
    ("GAF", "GAF"),
]

def canon_subject(x) -> str:
    import re
    s = str(x).strip()
    m = re.search(r"(\d+)", s)
    return f"NB_subject_{int(m.group(1))}" if m else s

def _stars(p: float) -> str:
    if p is None or (isinstance(p, float) and (not np.isfinite(p))):
        return ""
    if p < 1e-4: return "****"
    if p < 1e-3: return "***"
    if p < 1e-2: return "**"
    if p < 5e-2: return "*"
    return ""

def _fmt_q(q: float) -> str:
    if q is None or (isinstance(q, float) and (not np.isfinite(q))):
        return ""
    return f"{q:.4g}" + _stars(q)

def _find_template(root: Path, name: str) -> Path:
    for c in [
        root / "templates" / name,
        root / "template" / name,
        root / "supp_tables" / name,
        Path(__file__).resolve().parent / name,
        Path(__file__).resolve().parent.parent / name,
    ]:
        if c.exists():
            return c
    raise FileNotFoundError(name)

def write_xlsx(df: pd.DataFrame, template_path: Path, out_path: Path) -> None:
    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]

    for r in range(3, 250):
        for c in range(1, 8):
            ws.cell(r, c).value = None

    r0 = 3
    for _, row in df.iterrows():
        ws.cell(r0, 1).value = row["Outcome"]
        ws.cell(r0, 2).value = row["Predictor"]
        ws.cell(r0, 3).value = float(row["Effect"]) if np.isfinite(row["Effect"]) else None
        ws.cell(r0, 4).value = float(row["CI_low"]) if np.isfinite(row["CI_low"]) else None
        ws.cell(r0, 5).value = float(row["CI_high"]) if np.isfinite(row["CI_high"]) else None

        if ws.max_column >= 7:
            ws.cell(r0, 6).value = float(row["p_value"]) if np.isfinite(row["p_value"]) else None
            ws.cell(r0, 7).value = _fmt_q(float(row["q_fdr"])) if np.isfinite(row["q_fdr"]) else ""
        else:
            ws.cell(r0, 6).value = _fmt_q(float(row["q_fdr"])) if np.isfinite(row["q_fdr"]) else ""

        r0 += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def main():
    root = Path(__file__).resolve().parents[1]
    data_dir = root / "data"
    out_dir = root / "outputs" / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)

    clin_csv = data_dir / "fig5a_source_public.csv"
    id_map  = data_dir / "fig5a_id_map_private.csv"
    rate_csv = data_dir / "rate_epoch_subject_level.csv"

    for p in [clin_csv, id_map, rate_csv]:
        if not p.exists():
            raise FileNotFoundError(p)

    clin = pd.read_csv(clin_csv)
    num_cols = ["events_sum", "minutes_sum"] + COVARS + [c for c, _ in PRED_MAP]
    for c in num_cols:
        if c in clin.columns:
            clin[c] = pd.to_numeric(clin[c], errors="coerce")
    clin["site_public"] = clin["site_public"].astype(str)

    if "group" in clin.columns:
        clin["group"] = clin["group"].astype(str).str.upper().replace({"SC": "SZ"})
        clin = clin[clin["group"].isin(["SZ"])].copy()

    mp = pd.read_csv(id_map)
    need = {"subject", "anon_id"}
    if not need.issubset(mp.columns):
        raise ValueError(f"id map missing columns: {sorted(need - set(mp.columns))}")
    mp = mp.copy()
    mp["subject"] = mp["subject"].map(canon_subject)
    mp["anon_id"] = mp["anon_id"].astype(str)

    rate = pd.read_csv(rate_csv)
    need2 = {"subject", "freq", "n_epochs", "n_events", "recording_len_s", "n_events_in_epochs", "sum_epoch_dur_s"}
    miss = need2 - set(rate.columns)
    if miss:
        raise ValueError(f"rate_epoch_subject_level missing columns: {sorted(miss)}")

    rate = rate.copy()
    rate["subject"] = rate["subject"].map(canon_subject)
    rate["freq"] = pd.to_numeric(rate["freq"], errors="coerce").astype(int)
    for c in ["n_epochs", "n_events", "n_events_in_epochs", "sum_epoch_dur_s", "recording_len_s"]:
        rate[c] = pd.to_numeric(rate[c], errors="coerce")
    rate = rate[rate["freq"].isin(FREQS_USE)].copy()

    rate["n_events_outside_epochs"] = rate["n_events"] - rate["n_events_in_epochs"]

    agg = rate.groupby(["subject"], as_index=False).agg(
        clustered_epochs=("n_epochs", "sum"),
        outside_ripple_count=("n_events_outside_epochs", "sum"),
        in_events=("n_events_in_epochs", "sum"),
        in_time_s=("sum_epoch_dur_s", "sum"),
    )
    agg["mean_within_epoch_density"] = agg["in_events"] / agg["in_time_s"]
    agg.loc[(agg["in_time_s"] <= 0) | (~np.isfinite(agg["mean_within_epoch_density"])), "mean_within_epoch_density"] = np.nan

    agg = agg.merge(mp, on="subject", how="left")
    agg = agg.dropna(subset=["anon_id"]).copy()
    agg["anon_id"] = agg["anon_id"].astype(str)

    D = clin.merge(agg, on="anon_id", how="inner")

    outcomes = [
        ("clustered_epochs", "Number of high-rate ripple epochs", "count", "IRR"),
        ("outside_ripple_count", "Ripple events outside high-rate epochs", "count", "IRR"),
        ("mean_within_epoch_density", "Mean ripple density within high-rate epochs (events/s)", "cont", "beta"),
    ]

    fam = sm.families.NegativeBinomial(link=sm.families.links.Log())

    rows = []
    for out_col, out_label, out_type, effect_type in outcomes:
        tmp = []
        for pred_col, pred_label in PRED_MAP:
            if pred_col not in D.columns:
                continue

            if out_type == "count":
                cols_need = ["site_public", "minutes_sum", "events_sum", out_col, pred_col] + COVARS
            else:
                cols_need = ["site_public", "events_sum", out_col, pred_col] + COVARS

            X = D.replace([np.inf, -np.inf], np.nan).dropna(subset=cols_need).copy()
            if X.empty:
                continue

            if out_type == "count":
                X = X[X["minutes_sum"] > 0].copy()
                if X.empty:
                    continue
                formula = f"{out_col} ~ {pred_col} + events_sum + " + " + ".join(COVARS) + " + C(site_public)"
                fit = smf.glm(
                    formula=formula,
                    data=X,
                    family=fam,
                    offset=np.log(X["minutes_sum"].astype(float))
                ).fit(cov_type="HC3")
                beta = float(fit.params[pred_col])
                se = float(fit.bse[pred_col])
                p = float(fit.pvalues[pred_col])
                eff = float(np.exp(beta))
                ci_low = float(np.exp(beta - 1.96 * se))
                ci_high = float(np.exp(beta + 1.96 * se))
            else:
                formula = f"{out_col} ~ {pred_col} + events_sum + " + " + ".join(COVARS) + " + C(site_public)"
                fit = smf.ols(formula=formula, data=X).fit(cov_type="HC3")
                beta = float(fit.params[pred_col])
                se = float(fit.bse[pred_col])
                p = float(fit.pvalues[pred_col])
                eff = beta
                ci_low = beta - 1.96 * se
                ci_high = beta + 1.96 * se

            tmp.append(dict(
                Outcome=out_label,
                Predictor=pred_label,
                Effect=eff,
                Effect_type=effect_type,
                CI_low=ci_low,
                CI_high=ci_high,
                p_value=p,
            ))

        if tmp:
            df_tmp = pd.DataFrame(tmp)
            df_tmp["q_fdr"] = multipletests(df_tmp["p_value"].values, method="fdr_bh")[1]
            rows.append(df_tmp)

    if not rows:
        raise RuntimeError("No models were fit. Check predictors/inputs and NA filtering.")

    out = pd.concat(rows, ignore_index=True)

    out_csv = out_dir / "Table_S11_clinical_associations_with_cluster_metrics.csv"
    out.to_csv(out_csv, index=False)

    try:
        tpl = _find_template(root, "supp_table11.xlsx")
        out_xlsx = out_dir / "Supplementary_Table_S11_clinical_associations_with_cluster_metrics.xlsx"
        write_xlsx(out, tpl, out_xlsx)
    except Exception as e:
        print("[WARN] Could not write xlsx:", repr(e))

    print("[OK] Written:")
    print(" -", out_csv)

if __name__ == "__main__":
    main()
