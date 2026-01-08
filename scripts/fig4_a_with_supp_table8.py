#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fig4a: Raw transition share (HC vs SZ) + Supplementary Table S8 (template-based)

Input:
  NN_open_code/data/transition_share_master.csv
Required columns:
  subject, group, pair, pct

Outputs:
  outputs/figures/Fig4a_transition_share_HC_vs_SZ.pdf

Supplementary (Table S8):
  outputs/tables/Supplementary_Table_S8_transition_share.xlsx (if template found)
  outputs/tables/Table_S8_transition_share.csv  (always)

Stats:
  - permutation test on mean difference (SZ - HC), two-sided
  - BH-FDR across 5 pairs
"""

import argparse
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from statsmodels.stats.multitest import multipletests


PAIR_ORDER = ["HIP→HIP (same)", "HIP→HIP (opposite)", "HIP→CTX", "CTX→HIP", "CTX→CTX"]
PAIR_ALIASES = {
    "HIP→HIP(same)": "HIP→HIP (same)",
    "HIP→HIP (same)": "HIP→HIP (same)",
    "HIP→HIP(same hemisphere)": "HIP→HIP (same)",

    "HIP→HIP(opposite)": "HIP→HIP (opposite)",
    "HIP→HIP (opposite)": "HIP→HIP (opposite)",
    "HIP→HIP(opposite hemisphere)": "HIP→HIP (opposite)",

    "HIP→CTX": "HIP→CTX",
    "CTX→HIP": "CTX→HIP",
    "CTX→CTX": "CTX→CTX",
}

COLOR = {"HC": "#4C72B0", "SZ": "#C44E52"}  # blue/red


def norm_group(g: str) -> str:
    s = str(g).strip().upper()
    if s in ("SC", "SCHIZOPHRENIA"):
        return "SZ"
    if s in ("HEALTHY", "CONTROL"):
        return "HC"
    return s

def norm_pair(p: str) -> str:
    s = str(p).strip()
    return PAIR_ALIASES.get(s, s)

def mean_se(x):
    x = np.asarray(x, float)
    x = x[np.isfinite(x)]
    if x.size == 0:
        return np.nan, np.nan, 0
    m = float(np.mean(x))
    se = float(np.std(x, ddof=1) / np.sqrt(x.size)) if x.size >= 2 else np.nan
    return m, se, int(x.size)

def cohens_d(sz, hc):
    """Cohen's d for difference in means (SZ - HC), pooled SD."""
    x = np.asarray(hc, float); x = x[np.isfinite(x)]
    y = np.asarray(sz, float); y = y[np.isfinite(y)]
    if x.size < 2 or y.size < 2:
        return np.nan
    vx = x.var(ddof=1); vy = y.var(ddof=1)
    denom = (x.size + y.size - 2)
    if denom <= 0:
        return np.nan
    pooled = np.sqrt(((x.size - 1) * vx + (y.size - 1) * vy) / denom)
    if not np.isfinite(pooled) or pooled <= 0:
        return np.nan
    return float((y.mean() - x.mean()) / pooled)

def perm_test_two_sided(hc, sz, n_perm=10000, seed=0):
    """
    Two-sided permutation test for mean difference:
      obs = mean(SZ) - mean(HC)
    p = (count(|perm|>=|obs|)+1)/(n_perm+1)
    """
    rng = np.random.default_rng(seed)
    x = np.asarray(hc, float); x = x[np.isfinite(x)]
    y = np.asarray(sz, float); y = y[np.isfinite(y)]
    if x.size < 2 or y.size < 2:
        return np.nan, np.nan
    obs = float(np.mean(y) - np.mean(x))
    cat = np.concatenate([x, y])
    nx = x.size
    diffs = np.empty(n_perm, float)
    for i in range(n_perm):
        rng.shuffle(cat)
        diffs[i] = float(np.mean(cat[nx:]) - np.mean(cat[:nx]))
    p = float((np.sum(np.abs(diffs) >= abs(obs)) + 1) / (n_perm + 1))
    return obs, p


def _stars(p: float) -> str:
    if p is None or (isinstance(p, float) and (not np.isfinite(p))):
        return ""
    if p < 1e-4:
        return "****"
    if p < 1e-3:
        return "***"
    if p < 1e-2:
        return "**"
    if p < 5e-2:
        return "*"
    return ""

def _fmt_p(p: float) -> str:
    if p is None or (isinstance(p, float) and (not np.isfinite(p))):
        return ""
    if p < 1e-4:
        return "<0.0001"
    return f"{p:.4g}"

def _fmt_q_with_stars(q: float) -> str:
    if q is None or (isinstance(q, float) and (not np.isfinite(q))):
        return ""
    s = _fmt_p(float(q))
    return s + _stars(float(q))

def _find_template(root: Path, name: str) -> Path:
    candidates = [
        root / "templates" / name,
        root / "template" / name,
        root / "supp_tables" / name,
        root / "data" / name,
        Path(__file__).resolve().parent / name,
        Path(__file__).resolve().parent.parent / name,
        Path(__file__).resolve().parents[2] / name,
    ]
    for c in candidates:
        if c.exists():
            return c
    raise FileNotFoundError(f"Template not found: {name}")

def _write_s8_from_template(df_meanse: pd.DataFrame, df_test: pd.DataFrame, template_path: Path, out_path: Path) -> None:
    """
    Template layout observed in supp_table8.xlsx:
      Row 2 is header
      Starting row 3:
        - For each pair: one HC row that includes diff/d/p/q, followed by one SZ row with only mean/se.
    Columns A-H:
      A Transition Type
      B Group
      C Mean transition Share (%)
      D Standard Error (%)
      E Difference (SZ-HC)
      F Cohen's d
      G p-value
      H FDR p-value (string with stars)
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]

    # Clear existing data rows (keep headers). Clear rows 3..(3+2*len(PAIR_ORDER)+5)
    start_row = 3
    end_row = start_row + 2 * len(PAIR_ORDER) + 5
    for r in range(start_row, end_row + 1):
        for c in range(1, 9):
            ws.cell(r, c).value = None

    r = start_row
    for pair in PAIR_ORDER:
        # means/se for both groups
        sub_hc = df_meanse[(df_meanse["pair"] == pair) & (df_meanse["group"] == "HC")]
        sub_sz = df_meanse[(df_meanse["pair"] == pair) & (df_meanse["group"] == "SZ")]
        m_hc = float(sub_hc["mean"].iloc[0]) if len(sub_hc) else np.nan
        se_hc = float(sub_hc["se"].iloc[0]) if len(sub_hc) else np.nan
        m_sz = float(sub_sz["mean"].iloc[0]) if len(sub_sz) else np.nan
        se_sz = float(sub_sz["se"].iloc[0]) if len(sub_sz) else np.nan

        # stats
        hit = df_test[df_test["pair"] == pair]
        diff = float(hit["diff_SZ_minus_HC"].iloc[0]) if len(hit) else np.nan
        d = float(hit["cohens_d"].iloc[0]) if len(hit) else np.nan
        p = float(hit["p_value"].iloc[0]) if len(hit) else np.nan
        q = float(hit["q_fdr"].iloc[0]) if len(hit) else np.nan

        # HC row (with stats)
        ws.cell(r, 1).value = pair
        ws.cell(r, 2).value = "HC"
        ws.cell(r, 3).value = None if not np.isfinite(m_hc) else round(m_hc, 3)
        ws.cell(r, 4).value = None if not np.isfinite(se_hc) else round(se_hc, 3)
        ws.cell(r, 5).value = None if not np.isfinite(diff) else round(diff, 3)
        ws.cell(r, 6).value = None if not np.isfinite(d) else round(d, 3)
        ws.cell(r, 7).value = None if not np.isfinite(p) else float(_fmt_p(p))
        ws.cell(r, 8).value = _fmt_q_with_stars(q) if np.isfinite(q) else ""

        # SZ row (means only)
        ws.cell(r + 1, 1).value = None
        ws.cell(r + 1, 2).value = "SZ"
        ws.cell(r + 1, 3).value = None if not np.isfinite(m_sz) else round(m_sz, 3)
        ws.cell(r + 1, 4).value = None if not np.isfinite(se_sz) else round(se_sz, 3)
        # E-H blank
        r += 2

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default=None)
    ap.add_argument("--input", type=str, default=None,
                    help="default: <root>/data/transition_share_master.csv")
    ap.add_argument("--n-perm", type=int, default=10000)
    ap.add_argument("--seed", type=int, default=0)
    args = ap.parse_args()

    if args.root is not None:
        root = Path(args.root).expanduser().resolve()
    else:
        try:
            root = Path(__file__).resolve().parents[1]
        except NameError:
            root = Path.cwd().resolve()

    in_csv = Path(args.input).expanduser().resolve() if args.input else (root / "data" / "transition_share_master.csv")
    out_tab = root / "outputs" / "tables"
    out_fig = root / "outputs" / "figures"
    out_tab.mkdir(parents=True, exist_ok=True)
    out_fig.mkdir(parents=True, exist_ok=True)

    if not in_csv.exists():
        raise FileNotFoundError(f"Missing input: {in_csv}")

    df = pd.read_csv(in_csv)
    need = {"subject", "group", "pair", "pct"}
    miss = need - set(df.columns)
    if miss:
        raise ValueError(f"transition_share_master.csv missing columns: {sorted(miss)}")

    df = df.copy()
    df["group"] = df["group"].map(norm_group)
    df["pair"] = df["pair"].map(norm_pair)
    df["pct"] = pd.to_numeric(df["pct"], errors="coerce")

    df = df[df["group"].isin(["HC", "SZ"])].copy()
    df = df[df["pair"].isin(PAIR_ORDER)].copy()
    df = df.dropna(subset=["pct"])

    # ---- group mean ± SE (by pair) ----
    rows = []
    for pair in PAIR_ORDER:
        for grp in ["HC", "SZ"]:
            vals = df.loc[(df["pair"] == pair) & (df["group"] == grp), "pct"].to_numpy(float)
            m, se, n = mean_se(vals)
            rows.append(dict(pair=pair, group=grp, n=int(n), mean=m, se=se))
    df_meanse = pd.DataFrame(rows)
  
    # ---- permutation tests (pair-wise), then BH across 5 pairs ----
    trows = []
    for pair in PAIR_ORDER:
        hc = df.loc[(df["pair"] == pair) & (df["group"] == "HC"), "pct"].to_numpy(float)
        sz = df.loc[(df["pair"] == pair) & (df["group"] == "SZ"), "pct"].to_numpy(float)
        obs, p = perm_test_two_sided(hc, sz, n_perm=args.n_perm, seed=args.seed + abs(hash(pair)) % 10000)
        d = cohens_d(sz, hc)
        trows.append(dict(
            pair=pair,
            n_HC=int(np.isfinite(hc).sum()),
            n_SZ=int(np.isfinite(sz).sum()),
            diff_SZ_minus_HC=float(obs) if np.isfinite(obs) else np.nan,
            cohens_d=float(d) if np.isfinite(d) else np.nan,
            p_value=float(p) if np.isfinite(p) else np.nan
        ))
    df_test = pd.DataFrame(trows)
    df_test["q_fdr"] = np.nan
    m = df_test["p_value"].notna()
    if m.any():
        df_test.loc[m, "q_fdr"] = multipletests(df_test.loc[m, "p_value"].values, method="fdr_bh")[1]
  
    # ---- Supplementary Table S8 (template-based) ----
    df_s8 = df_meanse.merge(df_test, on="pair", how="left")
    df_s8.to_csv(out_tab / "Table_S8_transition_share.csv", index=False)

    try:
        tpl = _find_template(root, "supp_table8.xlsx")
        out_xlsx = out_tab / "Supplementary_Table_S8_transition_share.xlsx"
        _write_s8_from_template(df_meanse, df_test, tpl, out_xlsx)
    except Exception as e:
        print("[WARN] Supplementary Table S8 was not written:", repr(e))

    # ---- plot (single panel) ----
    x = np.arange(len(PAIR_ORDER))
    width = 0.36

    fig, ax = plt.subplots(figsize=(8.4, 4.8))

    for i, grp in enumerate(["HC", "SZ"]):
        sub = df_meanse[df_meanse["group"] == grp].set_index("pair").reindex(PAIR_ORDER)
        y = sub["mean"].to_numpy(float)
        e = sub["se"].to_numpy(float)
        ax.bar(x + (i - 0.5) * width, y, width=width, yerr=e, capsize=3,
               color=COLOR[grp], edgecolor="black", linewidth=0.5, label=grp)

    # stars based on q_fdr
    y_top = np.nanmax(df_meanse["mean"].to_numpy(float) + df_meanse["se"].to_numpy(float))
    y_top = float(y_top * 1.12) if np.isfinite(y_top) else 1.0
    ax.set_ylim(0, y_top)

    for j, pair in enumerate(PAIR_ORDER):
        hit = df_test[df_test["pair"] == pair]
        if len(hit) == 1:
            q = hit["q_fdr"].iloc[0]
            if pd.notna(q) and float(q) < 0.05:
                star = "***" if float(q) < 0.001 else ("**" if float(q) < 0.01 else "*")
                ax.text(j, y_top * 0.98, star, ha="center", va="top", fontsize=12, fontweight="bold")

    ax.set_xticks(x)
    ax.set_xticklabels(PAIR_ORDER, rotation=0)
    ax.set_ylabel("Raw transition share (%)")
    ax.set_title("Transition share (HC vs SZ)", loc="left")
    ax.grid(axis="y", alpha=0.25)
    ax.legend(frameon=False)

    out_pdf = out_fig / "Fig4a_transition_share_HC_vs_SZ.pdf"
    fig.tight_layout()
    fig.savefig(out_pdf, dpi=300, bbox_inches="tight")
    plt.close(fig)

    print("[OK] Written:")
    print(" -", out_pdf)
    print(" -", out_tab / "Table_S8_transition_share.csv")
    if (out_tab / "Supplementary_Table_S8_transition_share.xlsx").exists():
        print(" -", out_tab / "Supplementary_Table_S8_transition_share.xlsx")


if __name__ == "__main__":
    main()
