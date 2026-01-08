#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Table S1 + Fig.1 Panels A/C (counts & durations) with the SAME statistics

Input:
  - NN_open_code/data/df_clean_expanded.csv
    required columns:
      site, subject, group, cond, freq, event_count, mean_event_length

Outputs:
  Tables:
    - NN_open_code/outputs/tables/TableS1_ripple_event_counts.csv
    - NN_open_code/outputs/tables/TableS2_ripple_event_durations.csv  (optional but useful)
  Figures:
    - NN_open_code/outputs/figures/Fig1A_counts_{site}.pdf
    - NN_open_code/outputs/figures/Fig1C_durations_{site}.pdf

Statistics (Table S1-compatible):
  - Statistic: mean(HC) − mean(SZ)  (raw scale; NO log transform)
  - Permutation test: two-sided label shuffle
  - p-value: proportion(|perm| >= |obs|)   (no +1 correction; matches your table granularity)
  - Cohen's d: (mean(HC) − mean(SZ)) / pooled SD
  - FDR: BH across frequencies within each site × condition × outcome
"""

from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt

try:
    from statsmodels.stats.multitest import multipletests
except Exception as e:
    raise RuntimeError("statsmodels is required for BH-FDR in this script.") from e


# =========================
# Normalization (FIXED for Cortex)
# =========================

def norm_site(x: str) -> str:
    """
    Keep site as the raw lower-case label in data (e.g., gundai/kumasou),
    and later map to display name Gunma/Kumagaya for tables/figures.
    """
    return str(x).strip().lower()

def display_site(site_raw: str) -> str:
    s = str(site_raw).lower()
    if "gundai" in s or "gunma" in s:
        return "Gunma"
    if "kumasou" in s or "kumagaya" in s:
        return "Kumagaya"
    return site_raw

def norm_group(x: str) -> str:
    s = str(x).strip().upper()
    if s in ("HC", "CONTROL", "HEALTHY"):
        return "HC"
    if s in ("SZ", "SCHIZOPHRENIA", "SC"):
        return "SZ"
    return s

def norm_cond(x: str) -> str:
    """
    FORCE mapping to Hippocampus / Cortex (Extrahippocampal -> Cortex).
    Handles already-capitalized labels too.
    """
    s = str(x).strip()
    s0 = s.lower()

    # Hippocampus variants
    if s0 in ("hippocampus", "hippo", "hip"):
        return "Hippocampus"

    # Cortex / extrahippocampal variants
    if s0 in (
        "cortex", "cortical", "ctx",
        "extrahippocampal", "extrahippocampus", "extrahippocapus",
        "extrahippocampal (cortex)", "extrahippocampal cortex",
        "extrahippocampal_cortex",
    ):
        return "Cortex"

    # If the data already contains standard labels:
    if s in ("Hippocampus", "Cortex"):
        return s
    if s in ("Extrahippocampal",):  # <-- this is the most common reason Cortex disappears
        return "Cortex"

    return s


# =========================
# Permutation test + effect size
# =========================

def permutation_test_mean_diff(x_hc, y_sz, *, n_perm=10000, rng=None):
    """
    Two-sided permutation test for mean difference:
      stat = mean(HC) − mean(SZ)
    p = proportion(|perm_stats| >= |obs|)
    """
    x = np.asarray(x_hc, float)
    y = np.asarray(y_sz, float)
    x = x[np.isfinite(x)]
    y = y[np.isfinite(y)]

    if x.size < 2 or y.size < 2:
        return np.nan, np.nan

    if rng is None:
        rng = np.random.default_rng(0)

    obs = float(x.mean() - y.mean())
    pooled = np.concatenate([x, y])
    nx = x.size

    perm_stats = np.empty(n_perm, dtype=float)
    for i in range(n_perm):
        rng.shuffle(pooled)
        perm_stats[i] = pooled[:nx].mean() - pooled[nx:].mean()

    p = float(np.mean(np.abs(perm_stats) >= abs(obs)))
    return obs, p

def cohens_d(x_hc, y_sz):
    x = np.asarray(x_hc, float)
    y = np.asarray(y_sz, float)
    x = x[np.isfinite(x)]
    y = y[np.isfinite(y)]
    if x.size < 2 or y.size < 2:
        return np.nan

    mx, my = x.mean(), y.mean()
    vx, vy = x.var(ddof=1), y.var(ddof=1)
    df = x.size + y.size - 2
    if df <= 0:
        return np.nan

    pooled_sd = np.sqrt(((x.size - 1) * vx + (y.size - 1) * vy) / df)
    return (mx - my) / pooled_sd if pooled_sd > 0 else np.nan


# =========================
# BH-FDR helper
# =========================

def bh_fdr(pvals: np.ndarray) -> np.ndarray:
    p = np.asarray(pvals, float)
    if p.size == 0:
        return p
    return multipletests(p, method="fdr_bh")[1]


# =========================
# Table builders
# =========================

def build_table(df: pd.DataFrame, metric: str, n_perm=10000, seed=42) -> pd.DataFrame:
    """
    Returns a long table with:
      Site, Condition, Frequency (Hz),
      Statistic (perm test), Mean HC, Mean SZ, SD HC, SD SZ, Cohen's d, p-value, FDR corrected p-value
    FDR: BH across frequencies within each Site×Condition×metric (metric fixed here).
    """
    rng = np.random.default_rng(42)
    rows = []

    for (site, cond, freq), sub in df.groupby(["site", "cond", "freq"], dropna=False):
        hc = sub[sub["group"] == "HC"]
        sz = sub[sub["group"] == "SZ"]

        x = pd.to_numeric(hc[metric], errors="coerce").dropna().to_numpy(float)
        y = pd.to_numeric(sz[metric], errors="coerce").dropna().to_numpy(float)
        if x.size < 2 or y.size < 2:
            continue

        stat, p = permutation_test_mean_diff(x, y, n_perm=n_perm, rng=rng)
        d = cohens_d(x, y)

        rows.append(dict(
            Site=display_site(site),
            Condition=cond,
            **{"Frequency (Hz)": int(freq)},
            **{"Statistic (permutation test)": float(stat)},
            **{"Mean HC": float(x.mean())},
            **{"Mean SZ": float(y.mean())},
            **{"SD HC": float(x.std(ddof=1))},
            **{"SD SZ": float(y.std(ddof=1))},
            **{"Cohen’s d": float(d)},
            **{"p-value": float(p)},
        ))

    out = pd.DataFrame(rows)
    if out.empty:
        out["FDR corrected p-value"] = []
        return out

    # FDR across frequencies within each Site×Condition (metric is fixed per table)
    out["FDR corrected p-value"] = np.nan
    for (site, cond), g in out.groupby(["Site", "Condition"]):
        out.loc[g.index, "FDR corrected p-value"] = bh_fdr(g["p-value"].values)

    out = out.sort_values(["Site", "Condition", "Frequency (Hz)"]).reset_index(drop=True)
    return out


def format_p_with_stars(p: float, decimals: int = 4) -> str:
    """Format p-value with significance stars (* <0.05, ** <0.01)."""
    if p is None or (isinstance(p, float) and (np.isnan(p) or np.isinf(p))):
        return ""
    try:
        p_float = float(p)
    except Exception:
        return str(p)
    stars = ""
    if p_float < 0.01:
        stars = "**"
    elif p_float < 0.05:
        stars = "*"
    fmt = f"{{:.{decimals}f}}"
    return fmt.format(p_float) + stars


def write_table_s1_excel(df_table: pd.DataFrame, out_xlsx: Path) -> None:
    """
    Write Supplementary Table S1 (ripple event counts) as an Excel file
    matching the shared template layout (two-row grouped headers).
    """
    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1: blank (template has an empty first row)
    # Row 2: group headers
    headers_row2 = [
        "Site", "Condition", "Frequency (Hz)", "Statistic (permutation test)",
        "Mean", None, "Standard Deviation", None, "Cohen’s d", " p-value", "FDR correctied  p-value"
    ]
    for c, v in enumerate(headers_row2, start=1):
        ws.cell(row=2, column=c, value=v)

    # Row 3: subheaders for grouped columns
    headers_row3 = [None, None, None, None, "HC", "SZ", "HC", "SZ", None, None, None]
    for c, v in enumerate(headers_row3, start=1):
        ws.cell(row=3, column=c, value=v)

    # Merge grouped header cells (E2:F2, G2:H2)
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)
    ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)

    # Style headers
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in [2, 3]:
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.font = header_font
            cell.alignment = center

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18

    # Map columns from df_table
    # Expected df_table columns from build_table():
    # Site, Condition, Frequency (Hz), Statistic (perm test),
    # Mean HC, Mean SZ, SD HC, SD SZ, Cohen's d, p-value, FDR corrected p-value
    col_map = {
        "Site": "Site",
        "Condition": "Condition",
        "Frequency (Hz)": "Frequency (Hz)",
        "Statistic (permutation test)": "Statistic (permutation test)",
        "Mean_HC": "Mean HC",
        "Mean_SZ": "Mean SZ",
        "SD_HC": "SD HC",
        "SD_SZ": "SD SZ",
        "Cohen_d": "Cohen’s d",
            "Cohen_d_alt": "Cohen's d",
        "p": "p-value",
        "p_fdr": "FDR corrected p-value",
    }

    # Prepare data rows (start at row 4)
    r0 = 4
    for i, row in df_table.iterrows():
        ws.cell(r0+i, 1, row.get(col_map["Site"], ""))
        ws.cell(r0+i, 2, row.get(col_map["Condition"], ""))
        ws.cell(r0+i, 3, row.get(col_map["Frequency (Hz)"], np.nan))
        ws.cell(r0+i, 4, row.get(col_map["Statistic (permutation test)"], np.nan))

        ws.cell(r0+i, 5, row.get(col_map["Mean_HC"], np.nan))
        ws.cell(r0+i, 6, row.get(col_map["Mean_SZ"], np.nan))
        ws.cell(r0+i, 7, row.get(col_map["SD_HC"], np.nan))
        ws.cell(r0+i, 8, row.get(col_map["SD_SZ"], np.nan))

        ws.cell(r0+i, 9, row.get(col_map["Cohen_d"], row.get(col_map["Cohen_d_alt"], np.nan)))
        ws.cell(r0+i, 10, row.get(col_map["p"], np.nan))

        # Template uses a string with stars in the last column
        p_fdr_val = row.get(col_map["p_fdr"], np.nan)
        ws.cell(r0+i, 11, format_p_with_stars(p_fdr_val, decimals=4))

    # Number formats
    num_formats = {
        3: "0",          # Frequency
        4: "0.000",      # Statistic
        5: "0.000", 6: "0.000",  # Means
        7: "0.000", 8: "0.000",  # SDs
        9: "0.000",      # Cohen's d
        10: "0.0000",    # p-value
    }
    for r in range(r0, r0 + len(df_table)):
        for c, fmt in num_formats.items():
            ws.cell(r, c).number_format = fmt
        # Align data cells
        for c in range(1, 12):
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

    # Column widths (reasonable defaults)
    widths = [12, 14, 14, 26, 12, 12, 18, 12, 10, 10, 18]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(out_xlsx)

# =========================
# Figure builders (Fig1 A/C)
# =========================

def mean_se_by_freq(df: pd.DataFrame, metric: str) -> pd.DataFrame:
    """
    Compute mean±SE per (site, cond, freq, group).
    """
    rows = []
    for (site, cond, freq, group), sub in df.groupby(["site","cond","freq","group"], dropna=False):
        v = pd.to_numeric(sub[metric], errors="coerce").dropna().to_numpy(float)
        if v.size == 0:
            continue
        se = float(v.std(ddof=1) / np.sqrt(v.size)) if v.size >= 2 else np.nan
        rows.append(dict(site=site, cond=cond, freq=float(freq), group=group, mean=float(v.mean()), se=se))
    return pd.DataFrame(rows)

def stars_from_q(q: float) -> str:
    if not np.isfinite(q):
        return ""
    if q < 0.01:
        return "**"
    if q < 0.05:
        return "*"
    return ""

def plot_fig1_metric(df: pd.DataFrame, table: pd.DataFrame, metric: str, out_pdf: Path):
    """
    Creates a 2-panel figure per site:
      - left: Cortex
      - right: Hippocampus
    Bars: HC vs SZ by frequency, mean±SE
    Significance: annotate with stars based on FDR corrected p-value from the table.
    """
    # colors consistent with typical Fig1 (HC blue, SZ red)
    color = {"HC": "#4C72B0", "SZ": "#C44E52"}

    site_raw = df["site"].iloc[0]
    site_name = display_site(site_raw)

    fig, axes = plt.subplots(1, 2, figsize=(12, 4.8), sharey=False)
    cond_order = ["Cortex", "Hippocampus"]

    for ax, cond in zip(axes, cond_order):
        sub = df[df["cond"] == cond]
        if sub.empty:
            ax.axis("off")
            ax.set_title(f"{cond} (no data)")
            continue

        freqs = sorted(sub["freq"].unique())
        x = np.arange(len(freqs))
        w = 0.36

        for gi, group in enumerate(["HC", "SZ"]):
            means = []
            ses = []
            for f in freqs:
                v = pd.to_numeric(sub[(sub["freq"]==f) & (sub["group"]==group)][metric], errors="coerce").dropna().to_numpy(float)
                means.append(float(np.mean(v)) if v.size else np.nan)
                ses.append(float(np.std(v, ddof=1)/np.sqrt(v.size)) if v.size >= 2 else np.nan)

            offset = (-w/2) if gi == 0 else (w/2)
            ax.bar(x + offset, means, width=w, yerr=ses, capsize=3,
                   color=color[group], edgecolor="black", linewidth=0.6, label=group)

        # annotate stars using table's FDR corrected p-value
        tsub = table[(table["Site"] == site_name) & (table["Condition"] == cond)]
        for i, f in enumerate(freqs):
            hit = tsub[tsub["Frequency (Hz)"] == int(f)]
            if len(hit) == 1:
                q = float(hit["FDR corrected p-value"].iloc[0])
                st = stars_from_q(q)
                if st:
                    ymax = ax.get_ylim()[1]
                    ax.text(i, ymax*1.02, st, ha="center", va="bottom", fontsize=12, fontweight="bold")

        ax.set_xticks(x)
        ax.set_xticklabels([str(int(f)) for f in freqs])
        ax.set_xlabel("Frequency (Hz)")
        ax.set_title(f"{site_name} · {cond}")
        ax.grid(True, axis="y", alpha=0.25)

        if metric == "event_count":
            ax.set_ylabel("Ripple events (per 5-min)")
        else:
            ax.set_ylabel("Event duration (ms)")  # if your unit is ms

    handles, labels = axes[0].get_legend_handles_labels()
    if handles:
        fig.legend(handles, labels, loc="upper right", frameon=False)

    fig.suptitle("Fig.1 Panels A/C", y=1.02)
    fig.tight_layout()
    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_pdf, dpi=300, bbox_inches="tight")
    plt.close(fig)


# =========================
# Entry point
# =========================

def main():
    root = Path(__file__).resolve().parents[1]
    in_csv = root / "data" / "df_clean_expanded.csv"
    out_tables = root / "outputs" / "tables"
    out_figs = root / "outputs" / "figures"
    out_tables.mkdir(parents=True, exist_ok=True)
    out_figs.mkdir(parents=True, exist_ok=True)

    if not in_csv.exists():
        raise FileNotFoundError(f"Input not found: {in_csv}")

    df = pd.read_csv(in_csv)

    # Normalize
    df = df.copy()
    df["site"] = df["site"].map(norm_site)
    df["subject"] = df["subject"].astype(str).str.strip()
    df["group"] = df["group"].map(norm_group)
    df["cond"] = df["cond"].map(norm_cond)
    df["freq"] = pd.to_numeric(df["freq"], errors="coerce")
    df["event_count"] = pd.to_numeric(df["event_count"], errors="coerce")
    df["mean_event_length"] = pd.to_numeric(df["mean_event_length"], errors="coerce")

    # Filter to Table S1 target
    df = df[
        df["cond"].isin(["Hippocampus", "Cortex"]) &
        df["group"].isin(["HC", "SZ"])
    ].dropna(subset=["site","cond","freq"])

    # --- Table S1 (counts) ---
    table_s1 = build_table(df, metric="event_count", n_perm=10000, seed=42)
    out_s1 = out_tables / "TableS1_ripple_event_counts.csv"
    table_s1.to_csv(out_s1, index=False)

    

# --- Table S2 (durations; optional but often needed) ---
    table_s2 = build_table(df, metric="mean_event_length", n_perm=10000, seed=42)
    out_s2 = out_tables / "TableS2_ripple_event_durations.csv"
    table_s2.to_csv(out_s2, index=False)

    # --- Figures per site ---
    for site in sorted(df["site"].unique()):
        df_site = df[df["site"] == site].copy()
        if df_site.empty:
            continue

        # counts figure (Fig1A)
        out_pdf_counts = out_figs / f"Fig1a_counts_{display_site(site)}.pdf"
        plot_fig1_metric(df_site, table_s1, metric="event_count", out_pdf=out_pdf_counts)

        # durations figure (Fig1C)
        out_pdf_dur = out_figs / f"Fig1c_durations_{display_site(site)}.pdf"
        plot_fig1_metric(df_site, table_s2, metric="mean_event_length", out_pdf=out_pdf_dur)

    print("[OK] Written tables:")
    print(" -", out_s1)
    print(" -", out_s2)
    print("[OK] Written figures to:", out_figs)


if __name__ == "__main__":
    main()