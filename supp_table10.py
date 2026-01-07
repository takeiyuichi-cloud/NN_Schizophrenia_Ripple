#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Supplementary Table S10 (template-based):
Associations between total ripple load and clinical measures (SZ only)

This script fits Negative Binomial GLMs, one model per clinical predictor:
  events_sum ~ predictor + covariates + C(site_public)
  offset = log(minutes_sum)

IRR = exp(beta_predictor)

Inputs:
  <root>/data/fig5a_source_public.csv  (SAFE)
Required columns:
  - site_public, events_sum, minutes_sum
  - covariates: age, sex, JART, sleepiness_pre, antipsychotics
  - clinical measures (any subset is acceptable; missing ones are skipped):
      PANSS_positive, PANSS_negative, PANSS_general, GAF

Outputs (<root>/outputs/tables/):
  - Supplementary_Table_S10_total_ripple_load_vs_clinical.xlsx
  - Table_S10_total_ripple_load_vs_clinical.Table_S10_total_ripple_load_vs_clinical.csv
"""

from __future__ import annotations
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl
import statsmodels.formula.api as smf
import statsmodels.api as sm
from statsmodels.stats.multitest import multipletests

COVARS = ["age", "sex", "JART", "sleepiness_pre", "antipsychotics"]
PRED_MAP = [
    ("PANSS_positive", "PANSS POS"),
    ("PANSS_negative", "PANSS NEG"),
    ("PANSS_pasological",  "PANSS GEN"),
    ("GAF",            "GAF"),
]

def _stars(p: float) -> str:
    if p is None or (isinstance(p,float) and (not np.isfinite(p))):
        return ""
    if p < 1e-4: return "****"
    if p < 1e-3: return "***"
    if p < 1e-2: return "**"
    if p < 5e-2: return "*"
    return ""

def _fmt_q(q: float) -> str:
    if q is None or (isinstance(q,float) and (not np.isfinite(q))):
        return ""
    return f"{q:.4g}" + _stars(q)

def _find_template(root: Path, name: str) -> Path:
    for c in [root/"templates"/name, root/"template"/name, root/"supp_tables"/name,
              Path(__file__).resolve().parent/name, Path(__file__).resolve().parent.parent/name]:
        if c.exists(): return c
    raise FileNotFoundError(name)

def write_xlsx(df: pd.DataFrame, template_path: Path, out_path: Path) -> None:
    wb=openpyxl.load_workbook(template_path)
    ws=wb[wb.sheetnames[0]]
    # clear rows 3-20 cols A-F
    for r in range(3, 30):
        for c in range(1, 7):
            ws.cell(r,c).value=None
    r0=3
    for _,row in df.iterrows():
        ws.cell(r0,1).value=row["Predictor"]
        ws.cell(r0,2).value=float(row["IRR"])
        ws.cell(r0,3).value=float(row["CI_low"])
        ws.cell(r0,4).value=float(row["CI_high"])
        ws.cell(r0,5).value=float(row["p_value"])
        ws.cell(r0,6).value=_fmt_q(float(row["q_fdr"])) if np.isfinite(row["q_fdr"]) else ""
        r0+=1
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def main():
    root = Path(__file__).resolve().parents[1]
    in_csv = root/"data"/"fig5a_source_public.csv"
    out_dir = root/"outputs"/"tables"
    out_dir.mkdir(parents=True, exist_ok=True)

    if not in_csv.exists():
        raise FileNotFoundError(in_csv)

    df=pd.read_csv(in_csv)
    # coerce numeric
    need_base=["events_sum","minutes_sum"] + COVARS
    for c in need_base:
        df[c]=pd.to_numeric(df.get(c), errors="coerce")
    df["site_public"]=df["site_public"].astype(str)

    rows=[]
    fam=sm.families.NegativeBinomial(link=sm.families.links.Log())
    for pred_col, pred_label in PRED_MAP:
        if pred_col not in df.columns:
            continue
        D=df.copy()
        D[pred_col]=pd.to_numeric(D[pred_col], errors="coerce")
        cols_need=["site_public","events_sum","minutes_sum",pred_col]+COVARS
        D = (
    D.replace([np.inf, -np.inf], np.nan)
     .dropna(subset=cols_need)
)
        D=D[D["minutes_sum"]>0].copy()
        if D.empty: 
            continue

        formula = "events_sum ~ " + pred_col + " + " + " + ".join(COVARS) + " + C(site_public)"
        offset=np.log(D["minutes_sum"].astype(float))
        fit=smf.glm(formula=formula, data=D, family=fam, offset=offset).fit(cov_type="HC3")
        beta=float(fit.params[pred_col])
        se=float(fit.bse[pred_col])
        p=float(fit.pvalues[pred_col])
        irr=float(np.exp(beta))
        ci_low=float(np.exp(beta-1.96*se))
        ci_high=float(np.exp(beta+1.96*se))
        rows.append(dict(Predictor=pred_label, IRR=irr, CI_low=ci_low, CI_high=ci_high, p_value=p))

    if not rows:
        raise RuntimeError("No models were fit. Check that clinical columns exist in fig5a_source_public.csv.")

    out=pd.DataFrame(rows)
    out["q_fdr"]=multipletests(out["p_value"], method="fdr_bh")[1] if len(out)>0 else np.nan

    out_csv=out_dir/"Table_S10_total_ripple_load_vs_clinical.csv"
    out.to_csv(out_csv, index=False)

    try:
        tpl=_find_template(root, "supp_table10.xlsx")
        out_xlsx=out_dir/"Supplementary_Table_S10_total_ripple_load_vs_clinical.xlsx"
        write_xlsx(out, tpl, out_xlsx)
    except Exception as e:
        print("[WARN] Could not write xlsx:", repr(e))

    print("[OK] Written:")
    print(" -", out_csv)

if __name__=="__main__":
    main()
