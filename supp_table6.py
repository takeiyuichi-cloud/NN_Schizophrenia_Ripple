#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Supplementary Table S6 builder: Surrogate-excess measures of ripple clustering in schizophrenia
(debug-merged / surrogate-aware input)

方針
- Table S6 は surrogate-excess（real - surrogate_mean）が必要であり、surrogate列を含むテーブルを入力にする。
- 入力は high_rate_epoch_debug_merged.csv（または同等）を想定する。
- 群差: permutation test（差 = mean(SZ) - mean(HC)、両側）
- FDR: metricごとに周波数（5 bands）方向で BH-FDR

Inputs (auto-detected; first found is used):
  - <root>/outputs/tables/high_rate_epoch_debug_merged.csv
  - <root>/data/high_rate_epoch_debug_merged.csv
  - user-specified --debug-csv

Optional:
  - <root>/data/df_clean_expanded.csv (group map if debug has no group)
  - user-specified --group-csv

Template (optional):
  - supp_table6.xlsx

Outputs:
  - <root>/outputs/tables/Supplementary_Table_S6_surrogate_excess.csv
  - <root>/outputs/tables/Supplementary_Table_S6_surrogate_excess.xlsx (if template found)
"""

from __future__ import annotations

import argparse
from pathlib import Path
import re
import numpy as np
import pandas as pd
import openpyxl
from statsmodels.stats.multitest import multipletests


FREQ_ORDER = [80, 120, 160, 200, 240]
N_PERM_DEFAULT = 10000
SEED_DEFAULT = 0

GROUP_COL_CANDIDATES = ["group", "Group", "diagnosis", "Dx", "GROUP", "label", "group_label"]
SID_COL_CANDIDATES = ["S_ID", "sid", "SID"]


def canon_subject(x) -> str:
    s = str(x)
    m = re.search(r"(\d+)", s)
    if not m:
        return s.strip()
    return f"NB_subject_{int(m.group(1))}"

def normalize_group_labels(s: pd.Series) -> pd.Series:
    x = s.astype(str).str.strip().str.upper()
    x = x.replace({"SCHIZOPHRENIA": "SZ", "SCZ": "SZ", "SC": "SZ"})
    x = x.replace({"HEALTHY": "HC", "CONTROL": "HC"})
    return x

def find_first_existing(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    return None

def coerce_numeric(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")

def permutation_test_diff_means(x_hc: np.ndarray, x_sz: np.ndarray, n_perm: int, seed: int = 0) -> tuple[float, float]:
    rng = np.random.default_rng(seed)
    x_hc = np.asarray(x_hc, float); x_sz = np.asarray(x_sz, float)
    x_hc = x_hc[np.isfinite(x_hc)]
    x_sz = x_sz[np.isfinite(x_sz)]
    if x_hc.size < 2 or x_sz.size < 2:
        return (np.nan, np.nan)
    obs = float(np.mean(x_sz) - np.mean(x_hc))
    pooled = np.concatenate([x_hc, x_sz])
    n_hc = x_hc.size
    diffs = np.empty(n_perm, float)
    for i in range(n_perm):
        rng.shuffle(pooled)
        diffs[i] = float(np.mean(pooled[n_hc:]) - np.mean(pooled[:n_hc]))
    p = float(np.mean(np.abs(diffs) >= abs(obs)))
    p = max(p, 1.0 / float(n_perm))
    return obs, p

def cohens_d(x_hc: np.ndarray, x_sz: np.ndarray) -> float:
    x_hc = np.asarray(x_hc, float); x_sz = np.asarray(x_sz, float)
    x_hc = x_hc[np.isfinite(x_hc)]
    x_sz = x_sz[np.isfinite(x_sz)]
    if x_hc.size < 2 or x_sz.size < 2:
        return np.nan
    v1 = np.var(x_hc, ddof=1); v2 = np.var(x_sz, ddof=1)
    denom = max(1, (x_hc.size + x_sz.size - 2))
    pooled = ((x_hc.size - 1) * v1 + (x_sz.size - 1) * v2) / denom
    if pooled <= 0 or not np.isfinite(pooled):
        return np.nan
    return float((np.mean(x_sz) - np.mean(x_hc)) / np.sqrt(pooled))

def _stars(p: float) -> str:
    if p is None or (isinstance(p, float) and (not np.isfinite(p))):
        return ""
    if p < 1e-4:
        return "****"
    if p < 1e-3:
        return "***"
    if p < 1e-2:
        return "**"
    if p < 5e-2:
        return "*"
    return ""

def _fmt_q_with_stars(q: float) -> str:
    if q is None or (isinstance(q, float) and (not np.isfinite(q))):
        return ""
    s = f"{q:.4g}"
    return s + _stars(q)


# --- IO discovery
def find_root() -> Path:
    try:
        return Path(__file__).resolve().parents[1]
    except NameError:
        return Path.cwd().resolve()

def find_debug_csv(root: Path, user_path: str | None) -> Path:
    if user_path:
        p = Path(user_path).expanduser()
        if not p.is_absolute():
            p = (root / p).resolve()
        if not p.exists():
            raise FileNotFoundError(f"--debug-csv not found: {p}")
        return p
    candidates = [
        root / "outputs" / "tables" / "high_rate_epoch_debug_merged.csv",
        root / "data" / "high_rate_epoch_debug_merged.csv",
        root / "outputs" / "tables" / "Fig3_high_rate_epoch_debug_merged.csv",
    ]
    for c in candidates:
        if c.exists():
            return c
    raise FileNotFoundError("Could not auto-detect high_rate_epoch_debug_merged.csv. Provide --debug-csv.")

def find_template(root: Path, name: str) -> Path | None:
    candidates = [
        root / "templates" / name,
        root / "template" / name,
        root / "supp_tables" / name,
        root / "data" / name,
        Path(__file__).resolve().parent / name,
        Path(__file__).resolve().parent.parent / name,
        Path(__file__).resolve().parents[2] / name,
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


# --- group attachment
def add_group_from_sid(df: pd.DataFrame, sid_col: str) -> pd.DataFrame:
    out = df.copy()
    sid = pd.to_numeric(out[sid_col], errors="coerce")
    out["group"] = sid.map({1: "HC", 2: "SZ"})
    return out

def attach_group_info(df: pd.DataFrame, root: Path, group_csv: Path | None) -> pd.DataFrame:
    out = df.copy()

    gcol = find_first_existing(out, GROUP_COL_CANDIDATES)
    if gcol is not None:
        out["group"] = normalize_group_labels(out[gcol])
        return out

    sidcol = find_first_existing(out, SID_COL_CANDIDATES)
    if sidcol is not None:
        out = add_group_from_sid(out, sidcol)
        out["group"] = normalize_group_labels(out["group"])
        return out

    if group_csv is not None and group_csv.exists():
        gdf = pd.read_csv(group_csv).copy()
        if "subject" not in gdf.columns:
            id_like = find_first_existing(gdf, ["ID", "id", "Subject", "SUBJECT"])
            if id_like is None:
                raise ValueError(f"group_csv must have 'subject' or ID-like col. cols={gdf.columns.tolist()}")
            gdf["subject"] = gdf[id_like].map(canon_subject).astype(str)
        gdf["subject"] = gdf["subject"].map(canon_subject).astype(str)

        gcol2 = find_first_existing(gdf, GROUP_COL_CANDIDATES)
        sidcol2 = find_first_existing(gdf, SID_COL_CANDIDATES)
        if gcol2 is not None:
            gdf["group"] = normalize_group_labels(gdf[gcol2])
        elif sidcol2 is not None:
            gdf = add_group_from_sid(gdf, sidcol2)
            gdf["group"] = normalize_group_labels(gdf["group"])
        else:
            raise ValueError(f"group_csv has neither group nor S_ID. cols={gdf.columns.tolist()}")

        gdf = gdf[["subject", "group"]].drop_duplicates("subject")
        out = out.merge(gdf, on="subject", how="left")
        return out

    dfc = root / "data" / "df_clean_expanded.csv"
    if dfc.exists():
        gdf = pd.read_csv(dfc).copy()
        if {"subject", "group"}.issubset(gdf.columns):
            gdf["subject"] = gdf["subject"].map(canon_subject).astype(str)
            gdf["group"] = normalize_group_labels(gdf["group"])
            gdf = gdf[gdf["group"].isin(["HC", "SZ"])][["subject", "group"]].drop_duplicates("subject")
            out = out.merge(gdf, on="subject", how="left")
            return out

    return out


# --- metric selection for S6
# Each metric needs real and surrogate-mean columns
S6_METRICS = [
    {
        "label": "Excess clustered-epoch count",
        "real_candidates": ["n_epochs", "n_epochs_total", "n_high_rate_epochs", "n_cluster_epochs"],
        "surr_candidates": [
            "surr_n_epochs_mean",
            "surr_mean_n_epochs",
            "surr_n_epochs__mean",
            "surr_n_epochs_mean_poisson",
        ],
    },
    {
        "label": "Excess mean within-epoch peak ripple rate (events/s)",
        "real_candidates": [
            "mean_epoch_max_rate_hz",
            "mean_epoch_max_rate_hz_x",
            "mean_epoch_max_rate_hz_y",
            "max_epoch_rate_hz",
            "max_epoch_rate_hz_x",
            "max_epoch_rate_hz_y",
        ],
        "surr_candidates": [
            "surr_mean_epoch_max_rate_hz_mean",
            "surr_max_epoch_rate_hz_mean",
            "surr_median_epoch_max_rate_hz_mean",
        ],
    },
    {
        "label": "Excess median epoch duration (s)",
        "real_candidates": [
            "median_epoch_dur_s",
            "median_epoch_duration_s",
            "epoch_median_dur_s",
            "median_dur_s",
        ],
        "surr_candidates": [
            "surr_median_epoch_dur_s_mean",
            "surr_median_epoch_duration_s_mean",
            "surr_epoch_median_dur_s_mean",
            "surr_median_dur_mean",
            "surr_median_epoch_dur_mean",
        ],
    },
]
def build_surrogate_excess(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """
    Returns:
      df_ex: original df with added columns excess__{i}
      missing: list of missing metric labels (no real or surr col)
    """
    out = df.copy()
    missing = []
    for i, spec in enumerate(S6_METRICS, start=1):
        real_col = find_first_existing(out, spec["real_candidates"])
        surr_col = find_first_existing(out, spec["surr_candidates"])
        if real_col is None or surr_col is None:
            missing.append(f"{spec['label']} (real={real_col}, surr={surr_col})")
            continue
        out[f"excess__{i}"] = coerce_numeric(out[real_col]) - coerce_numeric(out[surr_col])
        out[f"_real__{i}"] = real_col
        out[f"_surr__{i}"] = surr_col
    return out, missing


def compute_s6_table(df: pd.DataFrame, n_perm: int, seed: int) -> pd.DataFrame:
    rows = []
    for i, spec in enumerate(S6_METRICS, start=1):
        ex_col = f"excess__{i}"
        if ex_col not in df.columns:
            continue
        for f in FREQ_ORDER:
            sub = df[df["freq"] == f].copy()
            g = normalize_group_labels(sub["group"])
            x = pd.to_numeric(sub[ex_col], errors="coerce")

            x_hc = x[g == "HC"].to_numpy(float)
            x_sz = x[g == "SZ"].to_numpy(float)

            n_hc = int(np.isfinite(x_hc).sum())
            n_sz = int(np.isfinite(x_sz).sum())

            mean_hc = float(np.nanmean(x_hc)) if n_hc > 0 else np.nan
            mean_sz = float(np.nanmean(x_sz)) if n_sz > 0 else np.nan
            sd_hc = float(np.nanstd(x_hc, ddof=1)) if n_hc >= 2 else (0.0 if n_hc == 1 else np.nan)
            sd_sz = float(np.nanstd(x_sz, ddof=1)) if n_sz >= 2 else (0.0 if n_sz == 1 else np.nan)

            diff, p = permutation_test_diff_means(x_hc, x_sz, n_perm=n_perm, seed=seed + int(f) + (abs(hash(spec["label"])) % 100000))
            d = cohens_d(x_hc, x_sz)

            rows.append(dict(
                metric_label=spec["label"],
                freq=int(f),
                n_HC=n_hc, n_SZ=n_sz,
                mean_excess_HC=mean_hc,
                mean_excess_SZ=mean_sz,
                sd_excess_HC=sd_hc,
                sd_excess_SZ=sd_sz,
                diff_excess_SZ_minus_HC=diff,
                cohens_d=d,
                p_perm=p
            ))
    out = pd.DataFrame(rows)
    if out.empty:
        out["q_fdr"] = []
        return out

    out["q_fdr"] = np.nan
    for m, subm in out.groupby("metric_label", dropna=False):
        mask = subm["p_perm"].notna()
        if mask.any():
            q = multipletests(subm.loc[mask, "p_perm"].values, method="fdr_bh")[1]
            out.loc[subm.index[mask], "q_fdr"] = q
    return out


def write_s6_from_template(df_s6: pd.DataFrame, template_path: Path, out_path: Path) -> None:
    """
    Template expected layout (typical):
      col A: Metric label (merged block)
      col B: Frequency
      col C-D: Mean excess (HC, SZ)
      col E-F: SD excess (HC, SZ)
      col G: Diff (SZ-HC)
      col H: Cohen's d
      col I: p
      col J: FDR q + stars (string)
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]

    # clear data area generously (rows 4..100, cols 1..10)
    for r in range(4, 120):
        for c in range(1, 11):
            ws.cell(r, c).value = None

    row = 4
    for spec in S6_METRICS:
        subm = df_s6[df_s6["metric_label"] == spec["label"]].set_index("freq").reindex(FREQ_ORDER).reset_index()
        for i, f in enumerate(FREQ_ORDER):
            rr = row + i
            ws.cell(rr, 1).value = spec["label"] if i == 0 else None
            ws.cell(rr, 2).value = int(f)

            hit = subm[subm["freq"] == int(f)]
            vals = hit.iloc[0].to_dict() if len(hit) >= 1 else {}
            def _set_num(cell, x, nd=3):
                if x is None or (isinstance(x, float) and (not np.isfinite(x))):
                    cell.value = None
                else:
                    cell.value = round(float(x), nd)

            _set_num(ws.cell(rr, 3), vals.get("mean_excess_HC"), 3)
            _set_num(ws.cell(rr, 4), vals.get("mean_excess_SZ"), 3)
            _set_num(ws.cell(rr, 5), vals.get("sd_excess_HC"), 3)
            _set_num(ws.cell(rr, 6), vals.get("sd_excess_SZ"), 3)
            _set_num(ws.cell(rr, 7), vals.get("diff_excess_SZ_minus_HC"), 3)
            _set_num(ws.cell(rr, 8), vals.get("cohens_d"), 3)
            _set_num(ws.cell(rr, 9), vals.get("p_perm"), 4)

            q = vals.get("q_fdr")
            ws.cell(rr, 10).value = _fmt_q_with_stars(float(q)) if q is not None and np.isfinite(q) else ""

        row += 5

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default=None)
    ap.add_argument("--debug-csv", type=str, default=None)
    ap.add_argument("--group-csv", type=str, default=None)
    ap.add_argument("--n-perm", type=int, default=N_PERM_DEFAULT)
    ap.add_argument("--seed", type=int, default=SEED_DEFAULT)
    args = ap.parse_args()

    root = Path(args.root).expanduser().resolve() if args.root else find_root()
    out_tab = root / "outputs" / "tables"
    out_tab.mkdir(parents=True, exist_ok=True)

    debug_csv = find_debug_csv(root, args.debug_csv)
    df = pd.read_csv(debug_csv)
    if "subject" not in df.columns or "freq" not in df.columns:
        raise ValueError(f"debug table must contain subject,freq. cols={df.columns.tolist()}")

    df = df.copy()
    df["subject"] = df["subject"].map(canon_subject).astype(str)
    df["freq"] = pd.to_numeric(df["freq"], errors="coerce")
    df = df[df["freq"].isin(FREQ_ORDER)].copy()

    group_csv = None
    if args.group_csv:
        p = Path(args.group_csv).expanduser()
        if not p.is_absolute():
            p = (root / p).resolve()
        group_csv = p
    df = attach_group_info(df, root, group_csv)

    if "group" not in df.columns:
        raise RuntimeError("Group labels not found/attached for S6. Provide --group-csv or include group in debug table/df_clean_expanded.csv.")
    df["group"] = normalize_group_labels(df["group"])
    df = df[df["group"].isin(["HC", "SZ"])].copy()

    df2, missing = build_surrogate_excess(df)
    if missing:
        print("[WARN] Some S6 metrics could not be built (missing real/surr columns):")
        for m in missing:
            print(" -", m)

    df_s6 = compute_s6_table(df2, n_perm=args.n_perm, seed=args.seed)
    out_csv = out_tab / "Table_S6_surrogate_excess.csv"
    df_s6.to_csv(out_csv, index=False)

    tpl6 = find_template(root, "supp_table6.xlsx")
    if tpl6 is not None:
        out_xlsx = out_tab / "Supplementary_Table_S6_surrogate_excess.xlsx"
        try:
            write_s6_from_template(df_s6, tpl6, out_xlsx)
        except Exception as e:
            print("[WARN] Failed to write S6 xlsx from template:", repr(e))
            out_xlsx = None
    else:
        out_xlsx = None

    print("[OK] Written:")
    print(" -", out_csv)
    if out_xlsx is not None:
        print(" -", out_xlsx)


if __name__ == "__main__":
    main()
